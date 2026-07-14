from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
from io import BytesIO
from datetime import datetime
import re
import time
import uuid
import threading
import logging
import os

try:
    from metadata_fetcher import (fetch_metadata, fetch_metadata_by_tt,
                                  fetch_person, fetch_game, fetch_brand,
                                  warm_upcoming)
except Exception:  # keep the app running even if the module is missing
    def fetch_metadata(title, is_movie=True):
        return {}

    def fetch_metadata_by_tt(tt, is_movie=True, title=""):
        return {}

    def fetch_person(name):
        return {}

    def fetch_game(name):
        return {}

    def fetch_brand(name):
        return {}

    def warm_upcoming():
        pass

import json
import base64

import types as _types

# Ops ingest templates (reference/*.xlsx) are the AUTHORITATIVE logic source
# for studios/networks/keywords/roll-ups; the inlined tables below remain as
# fallback when the template files are absent.
try:
    import reference_data as TREF
except Exception:
    TREF = None

# Four additional ingest schemas (Beauty / Beverages / Sports Teams / General)
# with category+sub-category auto-detection and dropdown validation rules.
# Fails soft: if the modules/JSON are missing the app behaves exactly as before.
try:
    import os as _os
    from titleforge_ingest_ext import (detect_schema as _tfx_detect,
                                       fill_category as _tfx_fill,
                                       build_branddef_row as _tfx_build,
                                       BRANDDEF_COLUMNS as _TFX_COLUMNS,
                                       SCHEMAS as _TFX_SCHEMAS,
                                       GENERAL_TITLE_CATEGORIES as _TFX_MASTER)
    from titleforge_validator import (load_rules as _tfx_load_rules,
                                      validate_row as _tfx_validate)
    _TFX_RULES = _tfx_load_rules(_os.path.join(
        _os.path.dirname(_os.path.abspath(__file__)),
        'titleforge_validation_rules.json'))
    TFX_OK = True
except Exception as _e:  # pragma: no cover
    logging.warning(f"titleforge ingest extension unavailable: {_e}")
    TFX_OK = False
    _TFX_MASTER = []
    _TFX_RULES = {}
    _TFX_COLUMNS = {}
    _TFX_SCHEMAS = {}


def _tref():
    return TREF if (TREF is not None and getattr(TREF, "LOADED", False)) else None

# ---- Reference tables (inlined so no separate file can be missed on deploy) ----
# distributor (raw from BOM/Wikipedia/Wikidata/IMDb/TMDB) -> LF network label
_NETWORK_LABEL = {
    "lionsgate": "Lionsgate / Summit", "summit entertainment": "Lionsgate / Summit",
    "lionsgate films": "Lionsgate / Summit", "lionsgate premiere": "Lionsgate / Summit",
    "columbia pictures": "Sony / Columbia",
    "sony pictures releasing": "Sony / Columbia", "sony pictures": "Sony / Columbia",
    "sony pictures entertainment": "Sony / Columbia", "sony pictures classics": "Sony Classics",
    "20th century fox": "20th Century Studios", "20th century studios": "20th Century Studios",
    "walt disney studios motion pictures": "Disney", "walt disney pictures": "Disney",
    "amazon mgm studios": "Amazon MGM Studios", "amazon studios": "Amazon MGM Studios",
    "warner bros.": "Warner Bros.", "warner bros. pictures": "Warner Bros.",
    "warner bros. discovery": "Warner Bros.",
    "neon rated": "Neon",
    "pbs": "PBS network", "public broadcasting service": "PBS network",
    "pbs distribution": "PBS network",
    "cineverse entertainment": "Cineverse", "cineverse corp.": "Cineverse",
}
_NETWORK_TO_COMPANIES = {
    "20th Century Studios": "Walt Disney Pictures", "Amazon MGM Studios": "Amazon Studios",
    "Disney": "Walt Disney Pictures", "Lionsgate / Summit": "Lionsgate", "Neon": "Neon",
    "Sony / Columbia": "Sony Pictures", "Sony Classics": "Sony Pictures",
    "Warner Bros.": "Warner Bros. Pictures",
}
_NETWORK_TO_YOUTUBE = {
    "20th Century Studios": "http://www.youtube.com/user/FoxMovies",
    "Amazon MGM Studios": "http://www.youtube.com/channel/UCf5CjDJvsFvtVIhkfmKAwAA",
    "Atlas Distribution": "http://www.youtube.com/channel/UCMLA_XtSbnfjXHL2An8zfGg",
    "Aura Entertainment": "http://www.youtube.com/@AuraEntFilms",
    "Big World Pictures": "http://www.youtube.com/channel/UCx1mHWMsCO96ungWSwS5Udg",
    "Blue Fox": "http://www.youtube.com/channel/UCmHYPCM_h8Tw9JkI3UnrCvA",
    "Cineverse": "http://www.youtube.com/@cineverse_ent",
    "Dark Sky Films": "http://www.youtube.com/user/dsf2006",
    "Disney": "http://www.youtube.com/@pixar",
    "Fathom Events": "http://www.youtube.com/user/FathomEvents",
    "Fin & Fur Films": "http://www.youtube.com/@finfurfilms/videos",
    "GKIDS": "http://www.youtube.com/user/GKIDStv",
    "Giant Pictures": "http://www.youtube.com/@GiantPictures",
    "Greenwich Entertainment": "http://www.youtube.com/channel/UCLFmfzQaJE_YlgXtnkr3e_Q",
    "IFC Films": "http://www.youtube.com/user/IFCFilmsTube",
    "Iconic Events": "http://www.youtube.com/@iconicreleasing",
    "Independent Film Company": "http://www.youtube.com/@IndependentFilmCompany",
    "Indican Pictures": "http://www.youtube.com/user/IndicanPictures",
    "Janus Films": "http://www.youtube.com/user/janusfilmsnyc",
    "Kani Releasing": "http://www.youtube.com/@kani-releasing",
    "Kino Lorber": "http://www.youtube.com/user/kinolorber",
    "Lionsgate / Summit": "http://www.youtube.com/user/LionsgateLIVE",
    "MUBI": "http://www.youtube.com/@mubi",
    "Magnolia": "http://www.youtube.com/user/MagnoliaPictures",
    "Neon": "http://www.youtube.com/channel/UCpy5dRhZd-JbZP4NsrnLt1w",
    "Oscilloscope Pictures": "http://www.youtube.com/user/oscopelabs",
    "PBS network": "http://www.youtube.com/@PBS",
    "Persimmon": "http://www.youtube.com/@persimmonpresents",
    "Roadside Attractions": "http://www.youtube.com/user/RoadsideFlix",
    "Row K Entertainment": "http://youtube.com/@rowkpresents",
    "Sandbox Films": "http://www.youtube.com/@sandboxdocs",
    "Sony / Columbia": "http://www.youtube.com/@sonypictures",
    "Sony Classics": "http://www.youtube.com/user/SonyPicturesClassics",
    "Strand Releasing": "http://www.youtube.com/user/StrandReleasing",
    "Sumerian Pictures": "http://www.youtube.com/@SumerianRecords",
    "Trafalgar Releasing": "http://www.youtube.com/channel/UC_0NZhyl9KH0aMWXRnAKM4g",
    "Warner Bros.": "http://www.youtube.com/@WarnerBros",
    "Watermelon Pictures": "http://www.youtube.com/@watermelonpicturesco",
    "Well Go USA": "http://www.youtube.com/user/wellgousa",
}
_NETWORK_TO_MANAGER = {
    "20th Century Studios": "Disney Insights & Analytics + Disney Theatrical Research + Disney Ad Sales",
    "Disney": "Disney Insights & Analytics + Disney Theatrical Research + Disney Ad Sales",
    "Amazon MGM Studios": "Amazon PV Enterprise", "Lionsgate / Summit": "Lionsgate",
    "Neon": "Neon", "Sony / Columbia": "Sony Enterprise", "Sony Classics": "Sony Enterprise",
    "Warner Bros.": "Warner Bros.",
}
_NETWORK_TO_SUBCATEGORY = {
    "Disney": "Release - Wide\nStudio - Major",
    "Warner Bros.": "Language Type - English\nRelease - Wide\nStudio - Major",
    "Sony / Columbia": "Release - Wide\nStudio - Independent",
    "Amazon MGM Studios": "Release - Wide\nStudio - Independent",
    "AMC Network": "Release - Wide\nStudio - Independent",
    "Neon": "Language Type - English\nRelease - Wide\nStudio - Independent",
    "Cineverse": "Release - Wide\nStudio - Independent",
}
# extra brand_set lines a DAR row carries when its network's parent company
# has corporate roll-ups (learned from the manual file)
_DAR_ROLLUPS = {
    "Walt Disney Pictures": ("The Walt Disney Company > Film Roll-up\n"
                             "The Walt Disney Company > Film + TV + Publishing Roll-up\n"
                             "The Walt Disney Company > Overall Roll-up"),
    "Warner Bros. Pictures": ("Warner Bros. Pictures Films\n"
                              "WarnerMedia > Film Roll-up\n"
                              "WarnerMedia > Film + TV + Publishing Roll-up\n"
                              "WarnerMedia > Overall Roll-up"),
}
_GENRE_FIX = {"Sci-Fi": "Sci Fi", "Science Fiction": "Sci Fi", "Film-Noir": "Film Noir", "Rom-Com": "Romance"}


def _ref_ci_get(mapping, key):
    if key is None:
        return None
    k = str(key).strip()
    if k in mapping:
        return mapping[k]
    kl = k.lower()
    for mk, mv in mapping.items():
        if mk.lower() == kl:
            return mv
    return None


def _ref_normalize_network(raw):
    if not raw:
        return raw
    return _ref_ci_get(_NETWORK_LABEL, raw) or str(raw).strip()


def _ref_normalize_genres(genre_multiline):
    if not genre_multiline:
        return genre_multiline, ""
    parts = [p.strip() for p in str(genre_multiline).split("\n") if p.strip()]
    fixed = [_GENRE_FIX.get(p, p) for p in parts]
    seen = set()
    uniq = [g for g in fixed if not (g in seen or seen.add(g))]
    return "\n".join(uniq), (uniq[0] if uniq else "")


REF = _types.SimpleNamespace(
    NETWORK_TO_MANAGER=_NETWORK_TO_MANAGER,
    normalize_network=_ref_normalize_network,
    companies_for=lambda n: _ref_ci_get(_NETWORK_TO_COMPANIES, n) or "",
    youtube_for=lambda n: _ref_ci_get(_NETWORK_TO_YOUTUBE, n) or "",
    subcategory_for=lambda n: _ref_ci_get(_NETWORK_TO_SUBCATEGORY, n) or "",
    dar_rollup_for=lambda c: _ref_ci_get(_DAR_ROLLUPS, c) or "",
    normalize_genres=_ref_normalize_genres,
)

try:
    from validator import validate_workbook, DEFAULT_RULES
except Exception:  # validator is optional; page still loads
    validate_workbook = None
    DEFAULT_RULES = {"rules": []}

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
logging.basicConfig(level=logging.INFO)

# Define ALL 42 columns in EXACT order (matches Test_Run.xlsx: A -> AP)
COLUMNS = [
    'record_type', 'brand_id', 'title', 'title_created_date', 'title_category',
    'title_sub_category', 'genre', 'primary_genre', 'iso_mic', 'stock_exchange',
    'ticker_symbol', 'companies', 'brand_set', 'composite_brand_set', 'active',
    'released_on', 'domestic_opening_weekend_box_office', 'domestic_opening_weekend_screens',
    'domestic_opening_weekend_rank', 'street_date', 'network', 'facebook_page',
    'facebook_verified', 'twitter_handle', 'twitter_verified', 'instagram_user',
    'youtube_channel_username', 'youtube_channel_company', 'tiktok_user', 'linkedin_page',
    'threads_page', 'pinterest_user_username', 'pinterest_board', 'wikipedia_page',
    'rottentomatoes', 'imdb_id', 'metacritic',
    'twitter_search_terms', 'instagram_business_hashtags', 'twitter_search_term_keywords',
    'url_managers', 'last_reviewed'
]

# Social-media / metadata columns that identify a "full schema" upload
SOCIAL_COLUMNS = [
    'facebook_page', 'facebook_verified', 'twitter_handle', 'twitter_verified',
    'instagram_user', 'youtube_channel_username', 'youtube_channel_company',
    'tiktok_user', 'linkedin_page', 'threads_page', 'pinterest_user_username',
    'pinterest_board',
]

# Fixed keyword tail used in twitter_search_term_keywords (derived from Test_Run)
_KEYWORDS = ('"all new" or episode or watch or tv or show or series or season or '
             'binge or stream or film or movie or premiere or screening or feature '
             'or trailer or teaser or theater or release')


def _alnum(s):
    """Lowercase and strip everything except letters/digits (for hashtags).
    Accented characters are transliterated (é -> e) rather than dropped, so
    brand names like L'Oréal hash to 'loreal' instead of 'loral'."""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s or "")).encode(
        "ascii", "ignore").decode("ascii")
    return re.sub(r'[^a-z0-9]', '', s.lower())


def _title_variants(clean_title):
    """Lowercase title variants used in youtube_channel_username lines.
    Titles with a colon get TWO lines: punctuation-stripped first, then the
    original (matches the manual Ops format, e.g. the PBS documentary case)."""
    tl = clean_title.lower()
    stripped = re.sub(r'\s*:\s*', ' ', tl)
    stripped = re.sub(r'\s+', ' ', stripped).strip()
    return [stripped, tl] if stripped != tl else [tl]


def build_youtube_username(company_channel, clean_title, own_channel=""):
    """youtube_channel_username lines:
      - the title's OWN channel URL alone on the first line (if it has one)
      - then '<network channel>|<title variant>' for each title variant
    """
    lines = []
    own = (own_channel or "").strip()
    if own and own != (company_channel or "").strip():
        lines.append(own)
    if company_channel:
        lines.extend(f"{company_channel}|{v}" for v in _title_variants(clean_title))
    return "\n".join(lines)


def generate_search_terms(clean_title, network, year, is_dar, twitter_handle="",
                          network_clause=""):
    """Generate twitter_search_terms (AL) and twitter_search_term_keywords (AN).
    When the title has its own @handle, it gets its own leading lines
    (matching the manual Ops pattern)."""
    label = "DAR" if is_dar else "Operations - Core Title"
    t_hash = _alnum(clean_title)
    n_hash = _alnum(network)

    # twitter_search_terms
    lines = []
    handle = (twitter_handle or "").strip().lstrip("@").lower()
    if handle:
        if is_dar:
            lines.append(f"@{handle}|DAR|DAR")
        else:
            lines.append(f"@{handle}|TV Ops|TV Ops")
            lines.append(f"@{handle}|Film Ops|Film Ops")
            lines.append(f"@{handle}|Operations - Core Title|Operations - Core Title")
    lines.append(f"#{t_hash}|{label}|{label}")
    if n_hash:
        lines.append(f"#{t_hash}{n_hash}|{label}|{label}")
    terms = "\n".join(lines)

    # twitter_search_term_keywords -- the per-studio clause from the ingest
    # template wins (lowercased, matching the export format); else generic
    inner = []
    if network_clause:
        inner.append(network_clause.lower())
    elif network:
        inner.append(f'"{network.lower()}" or @{n_hash} or #{n_hash}')
    if year:
        inner.append(f'"{year}"')
    inner.append(_KEYWORDS)
    clause = "(" + " or ".join(inner) + ")"
    kw = f'("{clean_title.lower()}") {clause}|{label}|{label}'
    if is_dar:
        kw += "|2021-01-01"
    return terms, kw


# network -> url_managers team (from reference_data, learned from the manual file).
URL_MANAGER_MAP = {}
if REF is not None:
    URL_MANAGER_MAP = {k.lower(): v for k, v in REF.NETWORK_TO_MANAGER.items()}


def _first_line(v):
    if not v:
        return ""
    return str(v).split("\n")[0].strip()


def _resolve_manager(row):
    for key in (row.get("network"), row.get("companies")):
        m = URL_MANAGER_MAP.get((str(key) or "").strip().lower())
        if m:
            return m
    return ""


# companies values for which url_managers is NOT generated
URL_MANAGER_SKIP_COMPANIES = {"unknown", "pristine brand"}


def generate_url_managers(row):
    """Build url_managers: one 'platform|value|manager' line per social present.
    Skips titles whose companies is 'Unknown' or 'Pristine Brand'.
    Returns '' if skipped or if no manager is resolvable.
    """
    companies = (str(row.get("companies")) or "").strip().lower()
    if companies in URL_MANAGER_SKIP_COMPANIES:
        return ""
    manager = _resolve_manager(row)
    if not manager:
        return ""
    entries = []
    fb = row.get("facebook_page"); ig = row.get("instagram_user")
    yt = _first_line(row.get("youtube_channel_company"))
    tk = row.get("tiktok_user"); tw = row.get("twitter_handle")
    if fb:
        entries.append(f"facebook|{fb}|{manager}")
    if ig:
        entries.append(f"instagram|{ig}|{manager}")
    if yt:
        entries.append(f"youtube|{yt}|{manager}")
    if tk:
        entries.append(f"tiktok|{tk}|{manager}")
    if tw:
        entries.append(f"twitter|http://twitter.com/{tw}|{manager}")
    return "\n".join(entries)


def _norm_bool(v):
    """Normalise a truthiness/string into the lowercase 'true'/'false' the feed expects."""
    if isinstance(v, bool):
        return "true" if v else "false"
    sv = str(v).strip().lower()
    if sv in ("false", "0", "no", "n", ""):
        return "false"
    return "true"


# ======================= TV Shows (BrandIngest schema) =======================
# TV rows export in the 39-column ApplyBrandDefinitionReport / BrandIngest
# format (learned from the manual Ops file), NOT the movie schema.
TV_COLUMNS = [
    'record_type', 'brand_id', 'title', 'title_category', 'title_sub_category',
    'genre', 'primary_genre', 'rovi_id', 'ticker_symbol', 'title_content_windows',
    'companies', 'brand_set', 'active', 'released_on', 'box_office', 'street_date',
    'gross_screen', 'opening_weekend_box_office', 'network', 'brand_listing_hidden',
    'facebook_page', 'twitter_handle', 'instagram_user', 'youtube_channel_username',
    'tiktok_user', 'tumblr_page', 'pinterest_user_username', 'pinterest_board',
    'wikipedia_page', 'rottentomatoes', 'imdb_id', 'metacritic',
    'facebook_search_terms', 'twitter_search_terms', 'instagram_search_terms',
    'tumblr_search_terms', 'twitter_search_term_keywords', 'youtube_search_terms',
    'reddit_search_terms',
]

# corporate roll-up brand_set blocks (appended to DAR rows)
_DISNEY_TV = ("The Walt Disney Company > Overall Roll-up\n"
              "The Walt Disney Company > TV Roll-up\n"
              "The Walt Disney Company > TV {kind} Roll-up\n"
              "The Walt Disney Company > TV + Publishing Roll-up\n"
              "The Walt Disney Company > Film + TV + Publishing Roll-up")
_NBCU_TV = ("NBCUniversal > Overall Roll-up\nNBCUniversal > TV Roll-up\n"
            "NBCUniversal > TV {kind} Roll-up\nNBCUniversal > TV + Publishing Roll-up\n"
            "NBCUniversal > Film + TV + Publishing Roll-up")

# per-network reference: ticker, parent company, YouTube channel(s), network
# tier, twitter keyword clause, corporate roll-ups, extra DAR brand sets
_TV_NETWORK = {
    "Netflix": dict(
        ticker="NFLX", companies="Netflix",
        yt=["https://www.youtube.com/user/NewOnNetflix"], tier="Streaming",
        clause='"Netflix" OR @netflix OR #Netflix OR #NowonNetflix',
        extras="Netflix - Emerging Titles\nPV Monthly - Emerging Titles"),
    "Paramount+": dict(
        ticker="VIA", companies="Viacom",
        yt=["https://www.youtube.com/channel/UCrRttZIypNTA1Mrfwo745Sg"], tier="Streaming",
        clause='"Paramount+" OR "Paramount Plus" OR @paramountplus OR #ParamountPlus',
        extras="Paramount+ - Emerging Titles\nPV Monthly - Emerging Titles"),
    "Amazon Prime Video": dict(
        ticker="AMZN", companies="Amazon Prime Video",
        yt=["https://www.youtube.com/user/amazonstudios",
            "https://www.youtube.com/@AmazonMGMStudios"], tier="Streaming",
        clause=('"prime video" or "amazon studios" or "amazon mgm studios" or '
                '@primevideo or @amazonmgmstudio or @primemovies OR #primevideo '
                'or #amazonmgmstudios or #primemovies or #amazonstudios'),
        extras=("Amazon Prime Video TV Network\nAmazon Prime Video - Emerging Titles\n"
                "PV Monthly - Emerging Titles")),
    "NBC": dict(
        ticker="CMCSA", companies="NBCU Research - Entertainment Networks",
        yt=["https://www.youtube.com/user/NBC"], tier="Broadcast",
        clause='"NBC" OR @nbc OR #NBC OR #NBCNetwork',
        corp=_NBCU_TV.format(kind="Broadcast")),
    "ABC": dict(
        ticker="DIS", companies="American Broadcasting Company",
        yt=["https://www.youtube.com/user/ABCNetwork"], tier="Broadcast",
        clause='"ABC" OR @ABCNetwork OR #ABC OR #ABCNetwork',
        corp=_DISNEY_TV.format(kind="Broadcast")),
    "National Geographic": dict(
        ticker="DIS", companies="National Geographic",
        yt=["https://www.youtube.com/user/NationalGeographic"], tier="Ad Supported Cable",
        clause=('"National Geographic Channel" OR "Nat Geo" OR "NatGeo" OR @NatGeoTV '
                'OR #NationalGeographicChannel OR #NatGeoTV OR #NatGeo'),
        corp=_DISNEY_TV.format(kind="Cable")),
    "FX": dict(
        ticker="DIS", companies="FX Network",
        yt=["https://www.youtube.com/user/FXNetworks"], tier="Ad Supported Cable",
        clause='"FX" OR @FXNetworks OR #FX OR #FXNetwork',
        corp=_DISNEY_TV.format(kind="Cable"), extras="FX All Brands Roll-Up"),
    "Bravo": dict(
        ticker="CMCSA", companies="Bravo!",
        yt=["https://www.youtube.com/user/VideoByBravo"], tier="Ad Supported Cable",
        clause='"Bravo" OR @BravoTV OR #BravoTV OR #Bravo',
        corp=_NBCU_TV.format(kind="Cable")),
    "Adult Swim": dict(
        ticker="T", companies="Adult Swim",
        yt=["https://www.youtube.com/user/adultswim"], tier="Ad Supported Cable",
        daypart="Other",  # late-night network
        clause='"Adult Swim" OR @adultswim OR #AdultSwim',
        corp=("WarnerMedia > Overall Roll-up\nWarnerMedia > TV Roll-up\n"
              "WarnerMedia > TV Cable Roll-up\nWarnerMedia > TV + Publishing Roll-up\n"
              "WarnerMedia > Film + TV + Publishing Roll-up\nAT&T Overall Roll-Up")),
    "Food Network": dict(
        ticker="DISCB", companies="Unknown",
        yt=["https://www.youtube.com/user/FoodNetworkTV"], tier="Ad Supported Cable",
        clause='"Food Network" OR @FoodNetwork OR #foodnetwork',
        corp=("Discovery > TV Roll-up\nDiscovery > TV + Publishing Roll-up\n"
              "Discovery > Film + TV + Publishing Roll-up")),
    "History": dict(
        ticker="", companies="A&E Television Networks",
        yt=["https://www.youtube.com/user/historychannel"], tier="Ad Supported Cable",
        clause=('"History Channel" OR "History Network" OR @HISTORY OR '
                '#historychannel OR #historytv'),
        corp=("A+E > TV Roll-up\nA+E > TV + Publishing Roll-up\n"
              "A+E > Film + TV + Publishing Roll-up")),
    "BritBox": dict(
        ticker="", companies="Unknown",
        yt=["https://www.youtube.com/channel/UC0yD7rYO26CAbkOx4rJkCzg"], tier="Streaming",
        clause='"BritBox" OR #BritBox OR @BritBox_US'),
    "Tubi": dict(
        ticker="", companies="Unknown",
        yt=["https://www.youtube.com/channel/UCNDsk0uhSlG1-br1Ex2Rgfg"], tier="Streaming",
        clause='"Tubi" OR #Tubi OR @tubi'),
    "Shudder": dict(
        ticker="", companies="Unknown",
        yt=["https://www.youtube.com/channel/UCcCCIrXmIJOYamxWqeJzI2Q"], tier="Streaming",
        clause='"Shudder" OR @Shudder OR #Shudder', extras="Shudder TV Network"),
    "Starz": dict(
        ticker="LGF.A", companies="Starz Entertainment",
        yt=["https://www.youtube.com/user/Starz"], tier="Premium Cable",
        clause='"Starz" OR @STARZ OR #Starz OR #StarzNetwork OR #StarzTV'),
    "Great American Family": dict(
        ticker="", companies="Unknown",
        yt=["https://www.youtube.com/channel/UCjIRUzJ-6-4nyX3GIsfwOOg"],
        tier="Ad Supported Cable",
        clause=('"Great American Family" OR "on GAF" OR @GAfamilyTV OR '
                '#greatamericanfamilychannel OR #GAFTV')),
}

_TV_KEYWORD_TAIL = ('"All New" OR Episode OR Watch OR tv OR Show OR Series OR season '
                    'OR binge OR Stream OR Film OR Movie OR Premiere OR Screening OR '
                    'Feature OR Trailer OR Teaser OR theater OR release')

# unscripted genres win primary_genre, in this priority order; otherwise
# scripted shows are bucketed Drama unless they are Comedy-without-Drama
_TV_UNSCRIPTED_PRIORITY = ["Game Show", "Reality", "Sport", "Documentary",
                           "Talk Show", "News"]


def _tv_net(network):
    return _ref_ci_get(_TV_NETWORK, network) or {}


def _tv_primary_genre(genres):
    for g in _TV_UNSCRIPTED_PRIORITY:
        if g in genres:
            return g
    if "Comedy" in genres and "Drama" not in genres:
        return "Comedy"
    return "Drama" if genres else ""


def _camel(s):
    """Hashtag form: keep case, '+' -> 'plus', drop everything non-alnum."""
    return re.sub(r'[^A-Za-z0-9]', '', str(s or '').replace('+', 'plus'))


def _https(u):
    return re.sub(r'^http://', 'https://', str(u or ''))


def _strip_disambiguator(title):
    """'Steps (Netflix)' -> 'Steps' (used in YouTube username lines)."""
    return re.sub(r'\s*\([^)]*\)\s*$', '', title).strip()


def _tv_youtube_username(channels, clean_title):
    """One '<channel>|<Title>' line per network channel; titles containing a
    colon get a second pass with the colon removed (matches the manual file:
    colon variant lines come AFTER the original lines)."""
    t = _strip_disambiguator(clean_title)
    variants = [t]
    stripped = re.sub(r'\s*:\s*', ' ', t)
    stripped = re.sub(r'\s+', ' ', stripped).strip()
    if stripped != t:
        variants.append(stripped)
    return "\n".join(f"{ch}|{v}" for v in variants for ch in channels)


def _tv_search_terms(title, network, is_dar):
    label = "DAR" if is_dar else "TV Ops"
    # base rows hashtag the disambiguator-stripped title; DAR rows keep the
    # full title (both patterns are consistent in the manual Ops file)
    t = _camel(title if is_dar else _strip_disambiguator(title))
    n = _camel(network)
    lines = [f"#{t}|{label}"]
    if n:
        lines.append(f"#{t}{n}|{label}")
    return "\n".join(lines)


def _tv_keywords_and_reddit(title, network, year, program_type, is_dar,
                            clause="", reddit_clause=""):
    """twitter_search_term_keywords + reddit_search_terms.
    Base rows of Specials use the short 'tonight' pattern; DAR rows and
    everything else use the full pattern (per the manual Ops file).
    `clause`/`reddit_clause` come from the ingest template when available."""
    info = _tv_net(network)
    title = _strip_disambiguator(title)
    tail = "DAR|DAR|2021-01-01" if is_dar else "Operations - Core Title|Operations - Core Title"
    if program_type == "Special" and not is_dar:
        inner = f'tonight OR watch OR tv OR show OR program OR "{network}"'
        kw = f'("{title}")({inner})|{tail}'
        r_inner = re.sub(r'\s+(?:OR|or)\s+', ' | ', inner)
    else:
        clause = clause or info.get("clause")
        if not clause and network:
            clause = f'"{network}" OR @{_camel(network).lower()} OR #{_camel(network)}'
        parts = [p for p in (clause, f'"{year}"' if year else "", _TV_KEYWORD_TAIL) if p]
        inner = " OR ".join(parts)
        kw = f'("{title}") ({inner})|{tail}'
        # reddit: template's ready-made ' | ' clause when present, else transform
        r_clause = reddit_clause or re.sub(r'\s+(?:OR|or)\s+', ' | ', clause or '')
        r_parts = [p for p in (r_clause, f'"{year}"' if year else "",
                               re.sub(r'\s+(?:OR|or)\s+', ' | ', _TV_KEYWORD_TAIL)) if p]
        r_inner = " | ".join(r_parts)
    reddit = f'("{title.replace("+", " ")}") ({r_inner.replace("+", " ")})'
    if is_dar:
        reddit += "|2021-01-01"
    return kw, reddit


def create_tv_row(title, network="", metadata=None):
    """Create a TV Shows row in the 39-column BrandIngest schema.
    Values present in `metadata` always win over computed defaults."""
    metadata = metadata or {}
    is_dar = " - DAR" in title
    clean_title = re.sub(r"\s*-\s*DAR\s*$", "", title, flags=re.IGNORECASE).strip()

    eff_network = (str(metadata.get('network') or network or '')).strip()
    info = _tv_net(eff_network)
    # network record from the Ops ingest template (authoritative when present)
    tinfo = _tref().tv_network(eff_network) if (_tref() and eff_network) else None

    # ---- title_sub_category (4 lines) ----
    ptype = str(metadata.get('program_type') or '').strip() or "Series"
    tier = ""
    if tinfo and tinfo.get("network_type"):
        tier = re.sub(r'^Network - ', '', tinfo["network_type"]).strip()
    tier = tier or info.get("tier") or "Streaming"
    daypart = info.get("daypart") or (
        "Prime Time" if tier in ("Broadcast", "Ad Supported Cable") else "Other")
    lang = str(metadata.get('original_language') or 'en').strip().lower()
    lang_line = "English" if lang in ("en", "english", "") else "Other"
    _sub = (f"Daypart - {daypart}\nProgram Type - {ptype}\n"
            f"Language Type - {lang_line}\nNetwork - {tier}")

    # ---- companies / ticker / brand_set ----
    if is_dar:
        companies = "Pristine Brand"
        n = eff_network
        if ptype in ("Series", "Mini-Series"):
            brand_set = (f"{n} -- Episodic + Roll-Up\n{n}-- Episodic Network Roll-Up\n"
                         "LF // TV Universe\nLF // TV // Episodic\n"
                         "LF // TV // Episodic Plus\nPristine DAR Brands")
        else:  # Special / TV Movie
            film_line = "LF // Film - Majors + Independents\n" if ptype == "TV Movie" else ""
            brand_set = ("LF // TV Universe\nLF // TV // Episodic Plus\n"
                         f"{n} -- Episodic + Roll-Up\n{film_line}Pristine DAR Brands")
        # conglomerate roll-up block from the ingest template; inline fallback
        conglom = (_tref().tv_conglomerate(eff_network) if _tref() else "") or \
            "\n".join(x for x in (info.get("corp"), info.get("extras")) if x)
        if conglom:
            brand_set += "\n" + conglom
    else:
        companies = (tinfo.get("company") if tinfo else "") or \
            info.get("companies") or "Unknown"
        brand_set = "Competitive View"

    # ---- genre / primary ----
    _genre = metadata.get('genre', '')
    if REF is not None and _genre:
        _genre, _ = REF.normalize_genres(_genre)
    _primary = str(metadata.get('primary_genre') or '').strip()
    if not _primary:
        genres_list = [g for g in str(_genre).split("\n") if g]
        # the template's Order-of-Operations mapping is authoritative
        _primary = (_tref().tv_primary_genre(genres_list) if _tref() else "") or \
            _tv_primary_genre(genres_list)

    # ---- youtube / search terms ----
    channels = list(info.get("yt") or [])
    if tinfo and tinfo.get("youtube"):
        # template channel wins; the cell may hold several channels
        # (newline-separated, e.g. Amazon Prime Video)
        channels = [c.strip() for c in str(tinfo["youtube"]).split("\n") if c.strip()]
    _yt = str(metadata.get('youtube_channel_username') or '').strip()
    if not _yt and channels:
        _yt = _tv_youtube_username(channels, clean_title)
    rel = str(metadata.get('released_on') or '')
    year = rel[:4] if rel[:4].isdigit() else ''
    gen_terms = _tv_search_terms(clean_title, eff_network, is_dar)
    gen_kw, gen_reddit = _tv_keywords_and_reddit(
        clean_title, eff_network, year, ptype, is_dar,
        clause=(tinfo.get("twitter_clause") if tinfo else ""),
        reddit_clause=(tinfo.get("reddit_clause") if tinfo else ""))

    def mv(key, default=''):
        v = metadata.get(key, '')
        return v if v not in (None, '') else default

    row = {
        'record_type': mv('record_type', 'INGESTED'),
        'brand_id': metadata.get('brand_id', ''),
        'title': title,
        'title_category': 'TV Shows',
        'title_sub_category': mv('title_sub_category', _sub),
        'genre': _genre,
        'primary_genre': _primary,
        'rovi_id': metadata.get('rovi_id', ''),
        'ticker_symbol': mv('ticker_symbol',
                            (tinfo.get("ticker") if tinfo else "") or info.get("ticker", "")),
        'title_content_windows': metadata.get('title_content_windows', ''),
        'companies': mv('companies', companies),
        'brand_set': mv('brand_set', brand_set),
        'active': mv('active', 't'),
        'released_on': metadata.get('released_on', ''),
        'box_office': metadata.get('box_office', ''),
        'street_date': metadata.get('street_date', ''),
        'gross_screen': metadata.get('gross_screen', ''),
        'opening_weekend_box_office': metadata.get('opening_weekend_box_office', ''),
        'network': eff_network,
        'brand_listing_hidden': mv('brand_listing_hidden', 'f'),
        # per the manual file, per-title social handles stay blank for TV
        # (coverage runs through the network accounts)
        'facebook_page': metadata.get('facebook_page', ''),
        'twitter_handle': metadata.get('twitter_handle', ''),
        'instagram_user': metadata.get('instagram_user', ''),
        'youtube_channel_username': _yt,
        'tiktok_user': metadata.get('tiktok_user', ''),
        'tumblr_page': metadata.get('tumblr_page', ''),
        'pinterest_user_username': metadata.get('pinterest_user_username', ''),
        'pinterest_board': metadata.get('pinterest_board', ''),
        'wikipedia_page': _https(metadata.get('wikipedia_page', '')),
        'rottentomatoes': _https(metadata.get('rottentomatoes', '')),
        'imdb_id': _https(metadata.get('imdb_id', '')),
        'metacritic': _https(metadata.get('metacritic', '')),
        'facebook_search_terms': metadata.get('facebook_search_terms', ''),
        'twitter_search_terms': mv('twitter_search_terms', gen_terms),
        'instagram_search_terms': metadata.get('instagram_search_terms', ''),
        'tumblr_search_terms': metadata.get('tumblr_search_terms', ''),
        'twitter_search_term_keywords': mv('twitter_search_term_keywords', gen_kw),
        'youtube_search_terms': metadata.get('youtube_search_terms', ''),
        'reddit_search_terms': mv('reddit_search_terms', gen_reddit),
    }
    return row


# ======================= Talent (BrandDef schema) =======================
# Talent rows export in the 38-column BrandDef format (from the Talent ingest
# template): ONE row per person, title suffixed ' - DAR', companies 'Pristine
# Brand', brand_set 'LF // Talent\nPristine DAR Brands', a single
# '#name|DAR|DAR' twitter search term, and a 3-line sub-category
# (Talent Subtype / Gender / Talent Type).
TALENT_COLUMNS = [
    'brand_id', 'title', 'title_category', 'title_sub_category', 'genre',
    'primary_genre', 'rovi_id', 'ticker_symbol', 'title_content_windows',
    'companies', 'brand_set', 'active', 'released_on', 'box_office',
    'street_date', 'gross_screen', 'opening_weekend_box_office', 'network',
    'facebook_page', 'twitter_handle', 'instagram_user',
    'youtube_channel_username', 'tiktok_user', 'linkedin_page', 'tumblr_page',
    'pinterest_user_username', 'pinterest_board', 'wikipedia_page',
    'rottentomatoes', 'imdb_id', 'metacritic', 'facebook_search_terms',
    'twitter_search_terms', 'instagram_search_terms', 'tumblr_search_terms',
    'twitter_search_term_keywords', 'youtube_search_terms', 'url_managers',
]

TALENT_DEFAULT_BRAND_SET = "LF // Talent\nPristine DAR Brands"

# discovered occupation keyword -> (Talent Type, subtype kind, subtype term)
# checked in order; 'Actor' becomes 'Actress' for Gender - Woman
_TALENT_OCC_MAP = [
    ('television presenter', 'Media Personality', 'Media Personality', 'TV'),
    ('television host', 'Media Personality', 'Media Personality', 'TV'),
    ('radio personality', 'Media Personality', 'Media Personality', 'Radio'),
    ('radio host', 'Media Personality', 'Media Personality', 'Radio'),
    ('podcaster', 'Media Personality', 'Media Personality', 'Podcaster'),
    ('youtuber', 'Internet Personality', 'Internet Personality', 'Content Creator'),
    ('internet celebrity', 'Internet Personality', 'Internet Personality', 'Influencer'),
    ('influencer', 'Internet Personality', 'Internet Personality', 'Influencer'),
    ('streamer', 'Internet Personality', 'Internet Personality', 'Streamer'),
    ('rapper', 'Musician', 'Musician', 'Rapper'),
    ('singer', 'Musician', 'Musician', 'Singer'),
    ('composer', 'Musician', 'Musician', 'Composer'),
    ('disc jockey', 'Musician', 'Musician', 'DJ / Producer'),
    ('record producer', 'Musician', 'Musician', 'DJ / Producer'),
    ('musician', 'Musician', '', ''),
    ('actor', 'Actor', '', ''),
    ('film director', 'Director', '', ''),
    ('director', 'Director', '', ''),
    ('comedian', 'Comedian', '', ''),
    ('politician', 'Politician', '', ''),
    ('journalist', 'Journalist', '', ''),
    ('chef', 'Chef', '', ''),
    ('model', 'Model', '', ''),
    ('dancer', 'Dancer', '', ''),
    ('choreographer', 'Dancer', '', ''),
    ('screenwriter', 'Writer', '', ''),
    ('author', 'Writer', '', ''),
    ('writer', 'Writer', '', ''),
    ('film producer', 'Producer', '', ''),
    ('producer', 'Producer', '', ''),
    ('activist', 'Activist', '', ''),
    ('entrepreneur', 'Entrepreneur', '', ''),
    ('businessperson', 'Entrepreneur', '', ''),
    ('scientist', 'Scientist', '', ''),
    ('photographer', 'Photographer', '', ''),
    ('fashion designer', 'Designer', '', ''),
    ('designer', 'Designer', '', ''),
    ('magician', 'Magician', '', ''),
    ('physician', 'Doctor', '', ''),
    ('teacher', 'Education', '', ''),
    ('professor', 'Education', '', ''),
    ('athlete', 'Athlete', '', ''),
]

# discovered sport label -> template subtype term (rest matched literally)
_TALENT_SPORT_ALIAS = {
    'association football': 'Soccer', 'american football': 'Football',
    'track and field': 'Running / Track & Field',
    'athletics': 'Running / Track & Field',
    'mixed martial arts': 'MMA', 'auto racing': 'Racing',
    'motorsport': 'Motorsports', 'professional wrestling': 'Pro Wrestling',
    'ice hockey': 'Ice Hockey', 'basketball': 'Basketball',
    'baseball': 'Baseball', 'tennis': 'Tennis', 'golf': 'Golf',
    'boxing': 'Boxing', 'swimming': 'Swimming', 'gymnastics': 'Gymnastics',
    'cricket': 'Cricket', 'surfing': 'Surfer', 'skateboarding': 'Skateboarding',
}


def _talent_classify(metadata):
    """(talent_type_line, subtype_line) from discovered occupations/sports."""
    occs = [str(o).lower() for o in (metadata.get('occupations') or [])]
    sports = [str(s).lower() for s in (metadata.get('sports') or [])]
    gender = str(metadata.get('gender') or '')
    ttype = subtype = ''
    if sports:
        ttype = 'Athlete'
        term = _TALENT_SPORT_ALIAS.get(sports[0], sports[0].title())
        if _tref():
            subtype = _tref().talent_subtype_for('Athlete', term)
        if not subtype:
            subtype = f"Talent Subtype - Athlete - {term}"
    else:
        for kw, typ, skind, sterm in _TALENT_OCC_MAP:
            if any(kw in o for o in occs):
                ttype = typ
                if skind and sterm:
                    subtype = (_tref().talent_subtype_for(skind, sterm)
                               if _tref() else '') or \
                        f"Talent Subtype - {skind} - {sterm}"
                break
    if ttype == 'Actor' and gender == 'Gender - Woman':
        ttype = 'Actress'
    if ttype == 'Politician':
        subtype = ('Talent Subtype - Politician - United States'
                   if metadata.get('us_citizen')
                   else 'Talent Subtype - Politician - International')
    return (f"Talent Type - {ttype}" if ttype else '', subtype)


def create_talent_row(title, metadata=None):
    """Create a Talent row in the 38-column BrandDef schema.
    Values present in `metadata` always win over computed defaults."""
    metadata = metadata or {}
    clean_name = re.sub(r"\s*-\s*DAR\s*$", "", title, flags=re.IGNORECASE).strip()
    out_title = f"{clean_name} - DAR"   # talent brands are DAR rows

    # sub-category: Subtype \n Gender \n Talent Type (template CONCAT order)
    _sub = str(metadata.get('title_sub_category') or '').strip()
    if not _sub:
        ttype_line, subtype_line = _talent_classify(metadata)
        if not ttype_line:
            ttype_line = 'Talent Type - Media Personality'  # template default
        gender_line = str(metadata.get('gender') or '').strip()
        _sub = "\n".join(x for x in (subtype_line, gender_line, ttype_line) if x)

    # twitter_search_terms: same logic as Movies/TV Shows (talent = DAR row)
    gen_terms, _ = generate_search_terms(
        clean_name, '', None, True,
        twitter_handle=str(metadata.get('twitter_handle') or ''))

    def mv(key, default=''):
        v = metadata.get(key, '')
        return v if v not in (None, '') else default

    row = {
        'brand_id': metadata.get('brand_id', ''),
        'title': out_title,
        'title_category': 'Talent',
        'title_sub_category': _sub,
        'genre': metadata.get('genre', ''),
        'primary_genre': metadata.get('primary_genre', ''),
        'rovi_id': metadata.get('rovi_id', ''),
        'ticker_symbol': metadata.get('ticker_symbol', ''),
        'title_content_windows': metadata.get('title_content_windows', ''),
        'companies': mv('companies', 'Pristine Brand'),
        'brand_set': mv('brand_set', TALENT_DEFAULT_BRAND_SET),
        'active': mv('active', 't'),
        'released_on': metadata.get('released_on', ''),
        'box_office': metadata.get('box_office', ''),
        'street_date': metadata.get('street_date', ''),
        'gross_screen': metadata.get('gross_screen', ''),
        'opening_weekend_box_office': metadata.get('opening_weekend_box_office', ''),
        'network': metadata.get('network', ''),
        'facebook_page': metadata.get('facebook_page', ''),
        'twitter_handle': metadata.get('twitter_handle', ''),
        'instagram_user': str(metadata.get('instagram_user') or '').lower(),
        'youtube_channel_username': metadata.get('youtube_channel_username', ''),
        'tiktok_user': metadata.get('tiktok_user', ''),
        'linkedin_page': metadata.get('linkedin_page', ''),
        'tumblr_page': metadata.get('tumblr_page', ''),
        'pinterest_user_username': metadata.get('pinterest_user_username', ''),
        'pinterest_board': metadata.get('pinterest_board', ''),
        'wikipedia_page': metadata.get('wikipedia_page', ''),
        'rottentomatoes': metadata.get('rottentomatoes', ''),
        'imdb_id': metadata.get('imdb_id', ''),
        'metacritic': metadata.get('metacritic', ''),
        'facebook_search_terms': metadata.get('facebook_search_terms', ''),
        'twitter_search_terms': mv('twitter_search_terms', gen_terms),
        'instagram_search_terms': metadata.get('instagram_search_terms', ''),
        'tumblr_search_terms': metadata.get('tumblr_search_terms', ''),
        'twitter_search_term_keywords': metadata.get('twitter_search_term_keywords', ''),
        'youtube_search_terms': metadata.get('youtube_search_terms', ''),
        'url_managers': metadata.get('url_managers', ''),
    }
    return row


# ===================== Video Games (BDR schema) =====================
# Games export in the 39-column Video-Game BDR format. Like movies/TV they
# get a base row (Operations - Core Title, brand_set 'Competitive View') and
# a ' - DAR' twin (DAR labels, brand_set 'LF // Video Games // Games').
GAME_COLUMNS = [
    'brand_id', 'title', 'title_category', 'title_sub_category', 'genre',
    'primary_genre', 'rovi_id', 'ticker_symbol', 'title_content_windows',
    'companies', 'brand_set', 'active', 'released_on', 'box_office',
    'street_date', 'gross_screen', 'opening_weekend_box_office', 'network',
    'brand_listing_hidden', 'facebook_page', 'twitter_handle', 'instagram_user',
    'youtube_channel_username', 'tiktok_user', 'tumblr_page',
    'pinterest_user_username', 'pinterest_board', 'wikipedia_page',
    'rottentomatoes', 'imdb_id', 'metacritic', 'facebook_search_terms',
    'twitter_search_terms', 'instagram_search_terms', 'tumblr_search_terms',
    'twitter_search_term_keywords', 'youtube_search_terms',
    'reddit_search_terms', 'url_managers',
]

GAME_DAR_BRAND_SET = "LF // Video Games // Games"
# fixed clause tails from the ingest template (incl. its 'Swtich 2' spelling)
_GAME_KW_TAIL = ('"Video Game" OR Playstation OR iOS OR PS4 OR PS5 OR Xbox OR '
                 'Switch OR Swtich 2 OR PC')
_GAME_RD_TAIL = ('"Video Game" | Playstation | iOS | PS4 | PS5 | Xbox | '
                 'Switch | Switch 2 | PC')

# discovered platform label -> template platform tail
_GAME_PLATFORM_ALIAS = {
    'playstation 5': 'PS5', 'playstation 4': 'PS4', 'playstation 2': 'PS2',
    'playstation': 'PS5', 'xbox series x': 'Xbox Series X',
    'xbox series x/s': 'Xbox Series X', 'xbox series s': 'Xbox Series X',
    'xbox one': 'Xbox One', 'nintendo switch 2': 'Switch 2',
    'nintendo switch': 'Switch', 'microsoft windows': 'PC', 'windows': 'PC',
    'macos': 'PC', 'linux': 'PC', 'ios': 'Mobile', 'android': 'Mobile',
    'game boy': 'Game Boy',
}


def _game_hashtag(name):
    """Template hashtag cleaning: case preserved; '&'->'and', '+'->'plus';
    removes space : , - ! ' . \\ ( )"""
    s = str(name or '')
    for a, b in ((' ', ''), (':', ''), (',', ''), ('-', ''), ('!', ''),
                 ("'", ''), ('.', ''), ('&', 'and'), ('+', 'plus'),
                 ('\\', ''), ('(', ''), (')', '')):
        s = s.replace(a, b)
    return '#' + s


def _game_platform_lines(platforms):
    out = []
    for p in (platforms or [])[:6]:
        pl = str(p).strip()
        tail = _GAME_PLATFORM_ALIAS.get(pl.lower())
        line = ''
        if _tref():
            line = _tref().game_platform_for(tail or pl)
        if not line:
            line = f"Platform - {tail or pl}"
        if line not in out:
            out.append(line)
    return out


def create_game_row(title, metadata=None):
    """Create a Video Game row in the 39-column BDR schema.
    Values present in `metadata` always win over computed defaults."""
    metadata = metadata or {}
    is_dar = " - DAR" in title
    clean_title = re.sub(r"\s*-\s*DAR\s*$", "", title, flags=re.IGNORECASE).strip()
    label = "DAR" if is_dar else "Operations - Core Title"

    developer = str(metadata.get('developer') or '').strip()
    dev_line = developer if developer.startswith('Developer - ') else \
        (f"Developer - {developer}" if developer else '')
    dev_name = dev_line.replace('Developer - ', '', 1)
    publisher = str(metadata.get('network') or '').strip()

    # sub-category = Developer line + up to 6 Platform lines
    _sub = str(metadata.get('title_sub_category') or '').strip()
    if not _sub:
        _sub = "\n".join(x for x in
                         [dev_line] + _game_platform_lines(metadata.get('platforms'))
                         if x)

    # genre (single, per template) + mapped primary
    _genre = str(metadata.get('genre') or '').split('\n')[0].strip()
    _primary = str(metadata.get('primary_genre') or '').strip()
    if not _primary and _genre:
        _primary = (_tref().game_primary_genre(_genre) if _tref() else '') or _genre

    # publisher YouTube channel ('...|' entries in the template) + title
    _yt = str(metadata.get('youtube_channel_username') or '').strip()
    if not _yt and publisher and _tref():
        pinfo = _tref().game_publisher(publisher)
        if pinfo and pinfo.get('youtube'):
            ch = pinfo['youtube']
            _yt = ch + clean_title if ch.endswith('|') else f"{ch}|{clean_title}"

    # developer keyword clauses (template tables; constructed fallback)
    dinfo = (_tref().game_developer(dev_name) if (_tref() and dev_name) else None) or {}
    tw_clause = dinfo.get('twitter_clause') or \
        (f'{_game_hashtag(dev_name)} OR "{dev_name}"' if dev_name else '"Video Game"')
    rd_clause = dinfo.get('reddit_clause') or \
        (f'{_game_hashtag(dev_name)} | "{dev_name}"' if dev_name else '"Video Game"')

    # twitter_search_terms: same logic as Movies/TV Shows (publisher plays
    # the network role for the '#<title><network>' line)
    gen_terms, _ = generate_search_terms(
        clean_title, publisher, None, is_dar,
        twitter_handle=str(metadata.get('twitter_handle') or ''))
    kw_tail = "|DAR|DAR|2021-01-01" if is_dar else \
        "|Operations - Core Title|Operations - Core Title"
    gen_kw = f'("{clean_title}") ({tw_clause} OR {_GAME_KW_TAIL}){kw_tail}'
    gen_reddit = f'("{clean_title}") ({rd_clause} | {_GAME_RD_TAIL})' + \
        ("|2021-01-01" if is_dar else "")

    def mv(key, default=''):
        v = metadata.get(key, '')
        return v if v not in (None, '') else default

    row = {
        'brand_id': metadata.get('brand_id', ''),
        'title': title,
        'title_category': 'Video Game',
        'title_sub_category': _sub,
        'genre': _genre,
        'primary_genre': _primary,
        'rovi_id': metadata.get('rovi_id', ''),
        'ticker_symbol': metadata.get('ticker_symbol', ''),
        'title_content_windows': metadata.get('title_content_windows', ''),
        'companies': mv('companies', 'Pristine Brand' if is_dar else 'Unknown'),
        'brand_set': mv('brand_set',
                        GAME_DAR_BRAND_SET if is_dar else 'Competitive View'),
        'active': mv('active', 't'),
        'released_on': metadata.get('released_on', ''),
        'box_office': metadata.get('box_office', ''),
        'street_date': metadata.get('street_date', ''),
        'gross_screen': metadata.get('gross_screen', ''),
        'opening_weekend_box_office': metadata.get('opening_weekend_box_office', ''),
        'network': publisher,
        'brand_listing_hidden': mv('brand_listing_hidden', 'f'),
        'facebook_page': metadata.get('facebook_page', ''),
        'twitter_handle': metadata.get('twitter_handle', ''),
        'instagram_user': str(metadata.get('instagram_user') or '').lower(),
        'youtube_channel_username': _yt,
        'tiktok_user': metadata.get('tiktok_user', ''),
        'tumblr_page': metadata.get('tumblr_page', ''),
        'pinterest_user_username': metadata.get('pinterest_user_username', ''),
        'pinterest_board': metadata.get('pinterest_board', ''),
        'wikipedia_page': metadata.get('wikipedia_page', ''),
        'rottentomatoes': metadata.get('rottentomatoes', ''),
        'imdb_id': metadata.get('imdb_id', ''),
        'metacritic': metadata.get('metacritic', ''),
        'facebook_search_terms': metadata.get('facebook_search_terms', ''),
        'twitter_search_terms': mv('twitter_search_terms', gen_terms),
        'instagram_search_terms': metadata.get('instagram_search_terms', ''),
        'tumblr_search_terms': metadata.get('tumblr_search_terms', ''),
        'twitter_search_term_keywords': mv('twitter_search_term_keywords', gen_kw),
        'youtube_search_terms': metadata.get('youtube_search_terms', ''),
        'reddit_search_terms': mv('reddit_search_terms', gen_reddit),
        'url_managers': metadata.get('url_managers', ''),
    }
    return row


def make_row(title, is_movie, network="", metadata=None, talent=False, game=False):
    """Dispatch: movies (42-col), TV (39-col BrandIngest), Talent (38-col
    BrandDef), Video Games (39-col BDR)."""
    if talent:
        return create_talent_row(title, metadata)
    if game:
        return create_game_row(title, metadata)
    if is_movie:
        return create_row(title, is_movie, network, metadata)
    return create_tv_row(title, network, metadata)


def create_row(title, is_movie, network="", metadata=None):
    """Create a data row for a title - ALL 42 COLUMNS POPULATED.

    Any value present in `metadata` overrides the computed default, so an
    uploaded row's channels (and every other field) are preserved.
    """
    metadata = metadata or {}
    is_dar = " - DAR" in title
    clean_title = re.sub(r"\s*-\s*DAR\s*$", "", title, flags=re.IGNORECASE).strip()
    # a trailing '(2026)' / '(Studio)' disambiguator stays in the title column
    # but is ignored for social fields and search terms
    social_title = _strip_disambiguator(clean_title)
    title_category = "Movies" if is_movie else "TV Shows"

    # Effective network = discovered/explicit network, else the passed arg;
    # normalise raw distributor -> LF network label (e.g. "Lionsgate" -> "Lionsgate / Summit").
    eff_network = (str(metadata.get('network') or network or '')).strip()
    if REF is not None and eff_network:
        eff_network = REF.normalize_network(eff_network)

    # studio record from the Ops ingest template (authoritative when present)
    sinfo = _tref().film_studio(eff_network) if (_tref() and eff_network) else None

    # sub-category is shared by the base title AND its DAR twin
    _sub_explicit = str(metadata.get('title_sub_category') or '').strip()
    _sub = _sub_explicit
    _scale = str(metadata.get('release_scale') or '').strip().title()
    if not _sub and sinfo:
        # template-driven: Language (when known) + Release scale + Studio Type
        lang = str(metadata.get('original_language') or '').strip().lower()
        lang_line = ''
        if lang:
            lang_line = 'Language Type - English\n' if lang in ('en', 'english') \
                else 'Language Type - Other\n'
        scale = _scale if _scale in ('Wide', 'Limited') else 'Limited'
        stype = sinfo.get('studio_type') or 'Studio - Independent'
        _sub = f"{lang_line}Release - {scale}\n{stype}"
    if not _sub and REF is not None:
        _sub = REF.subcategory_for(eff_network)
    if not _sub:
        _sub = 'Release - Limited\nStudio - Independent'
    # the upcoming-release-movies calendar knows the actual Wide/Limited scale;
    # it overrides the per-network default (but never an explicit upload value)
    if _scale in ('Wide', 'Limited') and not _sub_explicit:
        _sub = re.sub(r'Release - (Wide|Limited)', 'Release - ' + _scale, _sub)
    is_wide = 'release - wide' in _sub.lower()

    # curated PARENT company of the network (e.g. Warner Bros. -> Warner Bros. Pictures)
    parent = (sinfo.get('company') if sinfo else '') or \
        (REF.companies_for(eff_network) if (REF is not None and eff_network) else "")
    if parent == 'Unknown':
        parent = ''

    if is_dar:
        companies = "Pristine Brand"
        if is_movie:
            brand_set = "LF // Film - Majors + Independents\nPristine DAR Brands"
        else:
            brand_set = "Pristine DAR Brands"
        # major-studio DAR rows also carry the corporate roll-up brand sets
        # (per-studio block from the ingest template; inline map as fallback)
        rollup = ""
        if _tref():
            # the template's roll-up sheet may key by studio OR parent company
            rollup = _tref().film_rollup(eff_network) or _tref().film_rollup(parent)
        if not rollup and REF is not None and parent:
            rollup = REF.dar_rollup_for(parent)
        if rollup:
            brand_set += "\n" + rollup
    else:
        companies = parent or "Unknown"
        brand_set = "Competitive View"
        # Wide theatrical releases carry an extra brand_set line
        if is_wide:
            brand_set += "\n[Data Feed] Film - Wide Release + Custom Requests"

    # Release year for search-term generation
    rel = str(metadata.get('released_on') or metadata.get('title_created_date') or '')
    year = rel[:4] if rel[:4].isdigit() else ''

    gen_terms, gen_keywords = generate_search_terms(
        social_title, eff_network, year, is_dar,
        twitter_handle=str(metadata.get('twitter_handle') or ''),
        network_clause=(sinfo.get('twitter_clause') if sinfo else ''))

    # normalise genre tokens to the LF taxonomy (Sci-Fi -> Sci Fi, etc.)
    _genre = metadata.get('genre', '')
    _primary = metadata.get('primary_genre', '')
    if REF is not None and _genre:
        _genre, _primary_fix = REF.normalize_genres(_genre)
        if not _primary:            # keep a provided primary_genre; else derive
            _primary = _primary_fix

    # YouTube: company channel comes from the network; username lines combine
    # the title's own channel (if any) + '<network channel>|<title>' variants
    _yt_company = str(metadata.get('youtube_channel_company') or '').strip()
    if not _yt_company and sinfo and sinfo.get('youtube'):
        # movie schema uses http:// URLs; take the first channel if several
        _yt_company = re.sub(r'^https://', 'http://',
                             str(sinfo['youtube']).split('\n')[0].strip())
    if not _yt_company and REF is not None and eff_network:
        _yt_company = REF.youtube_for(eff_network)
    _yt_username = str(metadata.get('youtube_channel_username') or '').strip()
    if not _yt_username:
        _yt_username = build_youtube_username(
            _yt_company, social_title,
            own_channel=str(metadata.get('youtube_own_channel') or ''))

    def mv(key, default=''):
        """metadata value, falling back to default when missing OR blank."""
        v = metadata.get(key, '')
        return v if v not in (None, '') else default

    row = {
        'record_type': mv('record_type', 'INGESTED'),
        'brand_id': metadata.get('brand_id', ''),
        'title': title,
        'title_created_date': mv('title_created_date', datetime.now().strftime('%Y-%m-%d')),
        'title_category': mv('title_category', title_category),
        'title_sub_category': _sub,
        'genre': _genre,
        'primary_genre': _primary,
        'iso_mic': metadata.get('iso_mic', ''),
        'stock_exchange': metadata.get('stock_exchange', ''),
        'ticker_symbol': metadata.get('ticker_symbol', ''),
        'companies': mv('companies', companies),
        'brand_set': mv('brand_set', brand_set),
        'composite_brand_set': metadata.get('composite_brand_set', ''),
        'active': _norm_bool(metadata.get('active', True)),
        'released_on': metadata.get('released_on', ''),
        'domestic_opening_weekend_box_office': metadata.get('domestic_opening_weekend_box_office', ''),
        'domestic_opening_weekend_screens': metadata.get('domestic_opening_weekend_screens', ''),
        'domestic_opening_weekend_rank': metadata.get('domestic_opening_weekend_rank', ''),
        'street_date': metadata.get('street_date', ''),
        'network': eff_network,
        'facebook_page': metadata.get('facebook_page', ''),
        'facebook_verified': metadata.get('facebook_verified', ''),
        'twitter_handle': metadata.get('twitter_handle', ''),
        'twitter_verified': metadata.get('twitter_verified', ''),
        'instagram_user': metadata.get('instagram_user', ''),
        'youtube_channel_username': _yt_username,
        'youtube_channel_company': _yt_company,
        'tiktok_user': metadata.get('tiktok_user', ''),
        'linkedin_page': metadata.get('linkedin_page', ''),
        'threads_page': metadata.get('threads_page', ''),
        'pinterest_user_username': metadata.get('pinterest_user_username', ''),
        'pinterest_board': metadata.get('pinterest_board', ''),
        'wikipedia_page': metadata.get('wikipedia_page', ''),
        'rottentomatoes': metadata.get('rottentomatoes', ''),
        'imdb_id': metadata.get('imdb_id', ''),
        'metacritic': metadata.get('metacritic', ''),
        'twitter_search_terms': mv('twitter_search_terms', gen_terms),
        'instagram_business_hashtags': metadata.get('instagram_business_hashtags', ''),
        'twitter_search_term_keywords': mv('twitter_search_term_keywords', gen_keywords),
        'url_managers': metadata.get('url_managers', ''),
        'last_reviewed': metadata.get('last_reviewed', ''),
    }

    if not row.get('url_managers'):
        row['url_managers'] = generate_url_managers(row)

    return row


def _read_upload(src):
    """Read an uploaded CSV/XLSX into a DataFrame.
    `src` is a Werkzeug FileStorage OR a (bytes, filename) tuple (used by jobs)."""
    if isinstance(src, tuple):
        data, filename = src
        stream = BytesIO(data)
    else:
        filename = src.filename
        stream = src
    fn = (filename or '').lower()
    if fn.endswith('.csv'):
        df = pd.read_csv(stream)
    else:
        df = pd.read_excel(stream, engine='openpyxl')
    df.columns = [str(c).strip() for c in df.columns]
    df = df.where(pd.notnull(df), '')
    return df


def _merge_meta(base_meta, title, auto_fetch, is_movie=True):
    """Overlay auto-discovered metadata under any explicit metadata.
    Explicit values always win; auto-discovery only fills missing/blank fields.
    """
    if not auto_fetch:
        return base_meta or {}
    base = base_meta or {}
    tt = re.search(r"tt\d{5,}", str(base.get('imdb_id') or base.get('imdb_url') or ''))
    if tt:
        discovered = fetch_metadata_by_tt(tt.group(0), is_movie, title) or {}
    else:
        discovered = fetch_metadata(title, is_movie) or {}
    merged = dict(discovered)
    for k, v in base.items():
        if v not in (None, ''):
            merged[k] = v
    return merged


def _norm_kind(v, default='movie'):
    """'tvshow'/'TV Shows' -> 'tv'; 'Talent' -> 'talent'; 'Video Game(s)' ->
    'game'; Beauty/Beverages/Sports/General -> their tfx kind;
    else 'movie'. 'mixed' (or blank) falls back to the given default."""
    s = str(v or '').strip().lower()
    if 'talent' in s:
        return 'talent'
    if 'game' in s:
        return 'game'
    if 'beauty' in s:               # 'beauty', 'Health & Beauty'
        return 'beauty'
    if 'beverage' in s:             # 'beverages', 'Beverages'
        return 'beverages'
    if 'sport' in s:                # 'sports', 'Sports Franchise'
        return 'sports'
    if s == 'general':
        return 'general'
    if 'tv' in s:
        return 'tv'
    if s in ('', 'mixed', 'nan', 'none'):
        return default
    return 'movie'


# kinds handled by the titleforge extension (Beauty/Beverages/Sports/General)
TFX_KINDS = {'beauty', 'beverages', 'sports', 'general'}
_TFX_SHEET_LABELS = {'beauty': 'Beauty', 'beverages': 'Beverages',
                     'sports': 'Sports Teams', 'general': 'General'}


def create_tfx_row(title, kind, seed=None):
    """Build a BrandDef row for one of the four new schemas via the template
    logic in titleforge_ingest_ext. Explicit values from an uploaded row (seed)
    always win over derived ones, matching the app's other row builders."""
    seed = dict(seed or {})
    # drop blank/nan values so they don't mask derivation
    seed = {k: v for k, v in seed.items()
            if v is not None and str(v).strip() not in ('', 'nan', 'none', 'None')}
    seed.setdefault('Title', title)
    row = _tfx_build(kind, seed)
    row['title'] = row.get('title') or title
    # uploaded explicit values win
    cols = set(_TFX_COLUMNS.get(kind, []))
    low = {c.lower(): c for c in cols}
    for k, v in seed.items():
        col = low.get(str(k).strip().lower())
        if col:
            row[col] = v
    # twitter_search_terms: same logic as Movies/TV Shows (an explicit value
    # from the upload/payload still wins -- only the derived default changes)
    if not any(str(k).strip().lower() == 'twitter_search_terms' for k in seed):
        out_title = str(row.get('title') or title)
        base_title = re.sub(r"\s*-\s*DAR\s*$", "", out_title,
                            flags=re.IGNORECASE).strip()
        handle = str(row.get('twitter_handle') or '').strip()
        if 'twitter.com' in handle or 'x.com' in handle:
            handle = handle.rstrip('/').rsplit('/', 1)[-1]  # URL -> handle
        terms, _ = generate_search_terms(base_title,
                                         str(row.get('network') or ''),
                                         None, ' - DAR' in out_title,
                                         twitter_handle=handle)
        if terms:
            row['twitter_search_terms'] = terms
    row['_tfx_schema'] = kind   # internal routing marker; stripped on output
    return row


# ---------------- parallel auto-discovery (large-file support) ----------------
# Lookups used to run one title at a time; a 5,000-title file at ~2-5s per
# title could never finish. Discovery is I/O-bound, so a small thread pool
# gives a near-linear speed-up. Tune with the FETCH_WORKERS env var.
FETCH_WORKERS = max(1, int(os.getenv('FETCH_WORKERS', '8') or 8))


def _parallel_rows(items, worker, progress=None, parallel=True):
    """Run worker(index, item) -> [row, ...] for every item, preserving input
    order in the output. One failing title never kills the batch -- it just
    yields no rows (and is logged). progress(done, total) is thread-safe."""
    total = len(items)
    if not total:
        return []
    results = [None] * total
    state = {'done': 0}
    lock = threading.Lock()

    def _safe(i, item):
        try:
            results[i] = worker(i, item) or []
        except Exception as e:  # noqa: BLE001 -- fail soft per title
            logging.warning(f"title #{i + 1} failed during generation: {e}")
            results[i] = []
        finally:
            with lock:
                state['done'] += 1
                d = state['done']
            if progress:
                progress(d, total)

    if not parallel or total == 1 or FETCH_WORKERS == 1:
        for i, item in enumerate(items):
            _safe(i, item)
    else:
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=min(FETCH_WORKERS, total)) as ex:
            list(ex.map(lambda p: _safe(*p), enumerate(items)))

    out = []
    for r in results:
        out.extend(r or [])
    return out


def build_rows_from_upload(src, include_dar, auto_fetch=False, max_titles=None,
                           progress=None, default_kind='movie'):
    """Turn an uploaded file into fully-populated rows.

    max_titles caps titles processed BEFORE lookups (keeps Preview fast).
    progress(done, total) is called after each source title (for job progress).
    default_kind (movie/tv/talent) comes from the UI's Title-type selector and
    applies to rows that don't declare their own type/title_category.
    """
    df = _read_upload(src)
    lower_cols = {c.lower(): c for c in df.columns}
    has_full_schema = any(col in lower_cols for col in SOCIAL_COLUMNS + ['record_type', 'brand_id'])

    rows = []
    if has_full_schema:
        all_cols = list(dict.fromkeys(COLUMNS + TV_COLUMNS + TALENT_COLUMNS))
        rename_map = {lower_cols[c.lower()]: c for c in all_cols if c.lower() in lower_cols}
        df = df.rename(columns=rename_map)
        df = df.where(pd.notnull(df), '')
        records = df.to_dict('records')
        if max_titles:
            records = records[:max_titles]

        def _one_record(i, r):
            t = str(r.get('title', '')).strip()
            if not t:
                return []
            kind_r = _norm_kind(r.get('title_category'), default_kind)
            if kind_r in TFX_KINDS and TFX_OK:
                seed = r
                if auto_fetch:
                    # discovered socials/wikipedia fill blanks only --
                    # explicit values from the upload always win
                    seed = dict(fetch_brand(t) or {})
                    for k, v in r.items():
                        if v not in (None, ''):
                            seed[k] = v
                return [create_tfx_row(t, kind_r, seed)]
            if kind_r in ('talent', 'game'):
                if auto_fetch:
                    disc = dict((fetch_person(t) if kind_r == 'talent'
                                 else fetch_game(t)) or {})
                    for k, v in r.items():
                        if v not in (None, ''):
                            disc[k] = v
                    r = disc
                return [make_row(t, False, '', r,
                                 talent=(kind_r == 'talent'),
                                 game=(kind_r == 'game'))]
            is_movie_r = kind_r == 'movie'
            if auto_fetch:
                r = _merge_meta(r, t, True, is_movie=is_movie_r)
            # route through make_row so derived fields (network label, youtube
            # lines, brand sets, search terms) are computed consistently; explicit
            # values from the upload always win inside the row builders
            return [make_row(t, is_movie_r, str(r.get('network') or ''), r)]

        rows = _parallel_rows(records, _one_record, progress=progress,
                              parallel=auto_fetch)
    else:
        title_col = lower_cols.get('title') or df.columns[0]
        type_col = lower_cols.get('type') or lower_cols.get('title_category')
        network_col = lower_cols.get('network')
        specs = []
        for _, r in df.iterrows():
            title = str(r[title_col]).strip()
            if not title:
                continue
            if max_titles and len(specs) >= max_titles:
                break
            kind = _norm_kind(r[type_col] if type_col else '', default_kind)
            network = str(r[network_col]).strip() if network_col else ''
            specs.append((title, kind, network))
        def _one_spec(i, spec):
            title, kind, network = spec
            out = []
            if kind in TFX_KINDS and TFX_OK:
                seed = dict(fetch_brand(title) or {}) if auto_fetch else None
                out.append(create_tfx_row(title, kind, seed))
                if include_dar and ' - DAR' not in title:
                    out.append(create_tfx_row(f"{title} - DAR", kind, seed))
            elif kind == 'talent':
                meta = dict(fetch_person(title) or {}) if auto_fetch else {}
                out.append(make_row(title, False, '', meta, talent=True))
            elif kind == 'game':
                meta = dict(fetch_game(title) or {}) if auto_fetch else {}
                out.append(make_row(title, False, '', meta, game=True))
                if include_dar and ' - DAR' not in title:
                    out.append(make_row(f"{title} - DAR", False, '', meta, game=True))
            else:
                is_movie = kind == 'movie'
                meta = _merge_meta({}, title, auto_fetch, is_movie=is_movie)
                out.append(make_row(title, is_movie, network, meta))
                if include_dar and ' - DAR' not in title:
                    out.append(make_row(f"{title} - DAR", is_movie, network, meta))
            return out

        rows = _parallel_rows(specs, _one_spec, progress=progress,
                              parallel=auto_fetch)
    return rows


def build_rows_from_titles(data, max_titles=None, progress=None):
    """Build rows from a manual titles payload (JSON)."""
    titles = [t.strip() for t in data.get('titles', []) if t and t.strip()]
    if max_titles:
        titles = titles[:max_titles]
    include_dar = data.get('includeDar', True)
    auto_fetch = bool(data.get('autoFetch', False))
    def _one_title(i, title):
        kind = _norm_kind(data.get('titles_type', {}).get(title, 'movie'))
        network = data.get('networks', {}).get(title, '')
        base_meta = data.get('metadata', {}).get(title, {})
        out = []
        if kind in TFX_KINDS and TFX_OK:
            seed = base_meta
            if auto_fetch:
                # discovered socials/wikipedia fill blanks only --
                # explicit metadata from the payload always wins
                seed = dict(fetch_brand(title) or {})
                for k, v in (base_meta or {}).items():
                    if v not in (None, ''):
                        seed[k] = v
            out.append(create_tfx_row(title, kind, seed))
            if include_dar and ' - DAR' not in title:
                out.append(create_tfx_row(f"{title} - DAR", kind, seed))
        elif kind == 'talent':
            metadata = dict(fetch_person(title) or {}) if auto_fetch else {}
            for k, v in (base_meta or {}).items():
                if v not in (None, ''):
                    metadata[k] = v
            # talent = a single DAR row per person, no twin
            out.append(make_row(title, False, '', metadata, talent=True))
        elif kind == 'game':
            metadata = dict(fetch_game(title) or {}) if auto_fetch else {}
            for k, v in (base_meta or {}).items():
                if v not in (None, ''):
                    metadata[k] = v
            out.append(make_row(title, False, '', metadata, game=True))
            if include_dar and ' - DAR' not in title:
                out.append(make_row(f"{title} - DAR", False, '', metadata, game=True))
        else:
            is_movie = kind == 'movie'
            metadata = _merge_meta(base_meta, title, auto_fetch, is_movie=is_movie)
            out.append(make_row(title, is_movie, network, metadata))
            if include_dar and ' - DAR' not in title:
                out.append(make_row(f"{title} - DAR", is_movie, network, metadata))
        return out

    return _parallel_rows(titles, _one_title, progress=progress,
                          parallel=auto_fetch)


def _is_tv_row(r):
    return str(r.get('title_category', '')).lower() == 'tv shows'


def _is_talent_row(r):
    return str(r.get('title_category', '')).lower() == 'talent'


def _is_game_row(r):
    return 'game' in str(r.get('title_category', '')).lower()


def _rows_to_workbook(rows):
    """Write rows to an xlsx BytesIO. Movies use the 42-col schema, TV the
    39-col BrandIngest, Talent the 38-col BrandDef, Video Games the 39-col
    BDR; Beauty/Beverages/Sports/General use their template BrandDef layouts;
    mixed runs get one sheet per schema."""
    tfx = {}
    for r in rows:
        k = r.get('_tfx_schema')
        if k:
            tfx.setdefault(k, []).append(r)
    talent = [r for r in rows if not r.get('_tfx_schema') and _is_talent_row(r)]
    games = [r for r in rows if not r.get('_tfx_schema') and _is_game_row(r)]
    tv = [r for r in rows if not r.get('_tfx_schema') and _is_tv_row(r)]
    movies = [r for r in rows if not r.get('_tfx_schema')
              and not _is_tv_row(r) and not _is_talent_row(r) and not _is_game_row(r)]
    groups = [g for g in (movies, tv, talent, games, *tfx.values()) if g]
    if len(groups) > 1:
        sheets = []
        if movies:
            sheets.append(('Movies', movies, COLUMNS))
        if tv:
            sheets.append(('TV Shows', tv, TV_COLUMNS))
        if talent:
            sheets.append(('Talent', talent, TALENT_COLUMNS))
        if games:
            sheets.append(('Video Games', games, GAME_COLUMNS))
        for k, rws in tfx.items():
            sheets.append((_TFX_SHEET_LABELS.get(k, k.title()), rws,
                           _TFX_COLUMNS[k]))
    elif tfx:
        k, rws = next(iter(tfx.items()))
        sheets = [('BrandDef', rws, _TFX_COLUMNS[k])]
    elif talent:
        sheets = [('BrandDef', talent, TALENT_COLUMNS)]
    elif games:
        sheets = [('BDR', games, GAME_COLUMNS)]
    elif tv:
        sheets = [('BrandIngest', tv, TV_COLUMNS)]
    else:
        sheets = [('Sheet1', movies, COLUMNS)]
    out = BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as xw:
        for name, rws, cols in sheets:
            df = pd.DataFrame(rws).reindex(columns=cols)
            df = df.where(pd.notnull(df), '')
            df.to_excel(xw, sheet_name=name, index=False)
    out.seek(0)
    return out


# how many titles Preview samples (keeps auto-discovery fast on free tier)
PREVIEW_MAX_TITLES = 3


def collect_rows(preview=False):
    """Collect rows from either an uploaded file or a JSON titles payload.

    When preview=True only the first PREVIEW_MAX_TITLES titles are processed,
    BEFORE any auto-discovery, so the preview stays responsive.
    """
    max_titles = PREVIEW_MAX_TITLES if preview else None
    if request.files.get('file'):
        include_dar = request.form.get('includeDar', 'true').lower() != 'false'
        auto_fetch = request.form.get('autoFetch', 'false').lower() == 'true'
        default_kind = _norm_kind(request.form.get('titleType'))
        rows = build_rows_from_upload(request.files['file'], include_dar, auto_fetch,
                                      max_titles=max_titles,
                                      default_kind=default_kind)
    else:
        data = request.get_json(silent=True) or {}
        rows = build_rows_from_titles(data, max_titles=max_titles)
    return rows


@app.route('/')
def index():
    # wake the (free-tier) upcoming-release-movies service in the background
    # so the calendar index is ready by the time the user hits Generate
    warm_upcoming()
    return render_template('index.html')


@app.route('/api/lookup')
def api_lookup():
    """Debug helper: /api/lookup?title=Animal+Friends&type=movie[&tt=tt1234567]
    Shows exactly what auto-discovery finds for one title, plus the row that
    would be generated from it. Use this to verify enrichment after a deploy."""
    title = request.args.get('title', '').strip()
    tt = request.args.get('tt', '').strip()
    kind = request.args.get('type', 'movie').lower()
    if not (title or tt):
        return jsonify({'error': 'pass ?title= or ?tt='}), 400
    kind_n = _norm_kind(kind)
    if kind_n in TFX_KINDS and TFX_OK:
        meta = fetch_brand(title)
        row = create_tfx_row(title, kind_n, dict(meta))
        row.pop('_tfx_schema', None)
        return jsonify({'discovered': meta, 'row': row})
    if 'talent' in kind:
        meta = fetch_person(title)
        row = make_row(title, False, '', dict(meta), talent=True)
        return jsonify({'discovered': meta, 'row': row})
    if 'game' in kind:
        meta = fetch_game(title)
        row = make_row(title, False, '', dict(meta), game=True)
        return jsonify({'discovered': meta, 'row': row})
    is_movie = 'tv' not in kind
    if tt:
        meta = fetch_metadata_by_tt(tt, is_movie, title)
    else:
        meta = fetch_metadata(title, is_movie)
    row = make_row(title or tt, is_movie, '', dict(meta))
    return jsonify({'discovered': meta, 'row': row})


def _preview_payload(rows, preview_limited):
    """Shape enriched rows into the JSON the preview panel renders."""
    # preview shows the schema of the first title's category
    def _kind(r):
        return (r.get('_tfx_schema') or
                ('talent' if _is_talent_row(r) else
                 'game' if _is_game_row(r) else
                 'tv' if _is_tv_row(r) else 'movie'))
    first = _kind(rows[0])
    cols = {'talent': TALENT_COLUMNS, 'game': GAME_COLUMNS,
            'tv': TV_COLUMNS, 'movie': COLUMNS,
            **({k: v for k, v in _TFX_COLUMNS.items()} if TFX_OK else {})}[first]
    same = [r for r in rows if _kind(r) == first]
    df = pd.DataFrame(same)
    df = df.reindex(columns=cols).where(lambda x: pd.notnull(x), '')
    return {
        'total_rows': len(df),
        'preview': df.head(4).to_dict('records'),
        'columns': list(df.columns),
        'preview_limited': preview_limited,
    }


@app.route('/api/preview', methods=['POST'])
def preview_data():
    """Synchronous preview. Kept for backwards compatibility, but with
    Auto-discover ON the enrichment can outlive the HTTP request timeout
    (worker killed -> 502), so the UI uses /api/preview_async instead."""
    try:
        rows = collect_rows(preview=True)
        if not rows:
            return jsonify({'error': 'No titles provided'}), 400
        # figure out whether the source had more titles than we sampled
        if request.files.get('file'):
            preview_limited = True
        else:
            src = len((request.get_json(silent=True) or {}).get('titles', []))
            preview_limited = src > PREVIEW_MAX_TITLES
        return jsonify(_preview_payload(rows, preview_limited))
    except Exception as e:
        logging.error(f"Error previewing data: {str(e)}")
        return jsonify({'error': f"Error: {str(e)}"}), 500


@app.route('/api/generate', methods=['POST'])
def generate_excel():
    try:
        rows = collect_rows()
        if not rows:
            return jsonify({'error': 'No titles provided'}), 400
        output = _rows_to_workbook(rows)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Titles_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logging.error(f"Error generating Excel: {str(e)}")
        return jsonify({'error': f"Error: {str(e)}"}), 500


@app.route('/validator')
def validator_page():
    return render_template('validator.html',
                           default_rules=json.dumps(DEFAULT_RULES, indent=2))


@app.route('/api/validate', methods=['POST'])
def api_validate():
    try:
        if not request.files.get('file'):
            return jsonify({'error': 'Please upload a workbook (.xlsx or .csv).'}), 400
        if validate_workbook is None:
            return jsonify({'error': 'Validator module unavailable.'}), 500

        raw = request.form.get('rules')
        if request.files.get('rulesFile'):
            raw = request.files['rulesFile'].read().decode('utf-8', errors='replace')
        rules = None
        if raw and raw.strip():
            try:
                rules = json.loads(raw)
            except Exception as e:
                return jsonify({'error': f'Invalid rules JSON: {e}'}), 400

        xlsx_bytes, summary = validate_workbook(request.files['file'], rules)
        return jsonify({
            'summary': summary,
            'file_b64': base64.b64encode(xlsx_bytes).decode('ascii'),
            'filename': f"Validated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        })
    except Exception as e:
        logging.error(f"Error validating workbook: {str(e)}")
        return jsonify({'error': f"Error: {str(e)}"}), 500


# ========================= manual-file review ============================
# Reviews a manually prepared workbook against what the tool would generate
# (ingest-template logic + auto-discovery) and returns a highlighted copy
# with a Findings sheet (gaps + mismatches + suggested values) and a Summary.

# columns that are inherently manual / not derivable -> never flagged
REVIEW_SKIP_COLS = {
    'record_type', 'brand_id', 'title', 'title_category', 'title_created_date',
    'active', 'brand_listing_hidden', 'last_reviewed', 'rovi_id',
    'title_content_windows', 'composite_brand_set', 'iso_mic', 'stock_exchange',
    'box_office', 'street_date', 'gross_screen', 'opening_weekend_box_office',
    'domestic_opening_weekend_box_office', 'domestic_opening_weekend_screens',
    'domestic_opening_weekend_rank', 'facebook_verified', 'twitter_verified',
}


def _review_norm(v):
    """Comparison form: trimmed lines, scheme-insensitive URLs, no blanks."""
    s = str(v if v is not None else '').strip()
    if s.lower() in ('nan', 'none'):
        return ''
    s = s.replace('\r\n', '\n')
    s = '\n'.join(ln.strip() for ln in s.split('\n') if ln.strip())
    return re.sub(r'^https://', 'http://', s, flags=re.M)


def _sub_parts(sub):
    return {l.split(' - ', 1)[0].strip(): l.split(' - ', 1)[1].strip()
            for l in str(sub or '').split('\n') if ' - ' in l}


# ============ column-aware review rules (reviewer feedback, Jul 2026) ============
# Encodes the reviewed-file comments so curated values that are valid
# alternatives are no longer flagged as Mismatch.

_FANPAGE_RE = re.compile(r'fan[\s_-]?(page|club)|fanpage', re.I)
# never valid in a manual value (reviewer: /p/ and /people/ URLs are not valid)
_BAD_FB_MANUAL_RE = re.compile(r'facebook\.com/(p/|people/)', re.I)
# additionally never offered as a suggestion (unhelpful discovered URLs)
_BAD_FB_SUGG_RE = re.compile(
    r'facebook\.com/(p/|people/|profile\.php|pages/|\d+/*$)', re.I)


def _review_lines_ci(v):
    """Case-folded set of lines; any '||suffix' (added-date etc.) stripped."""
    out = set()
    for ln in str(v or '').split('\n'):
        ln = ln.strip()
        if ln:
            out.add(ln.split('||', 1)[0].strip().lower())
    return out


def _review_slug(s):
    return re.sub(r'[^a-z0-9]+', '', str(s or '').lower())


def _review_compare(col, manual_raw, expected_raw, title='', cat='',
                    manual_genre='', sub_raw=''):
    """Column-aware comparison. Returns (ok, suggested_str).

    Rules from the reviewed-file feedback:
      * title_sub_category / brand_set  - extra values are fine; only flag
        when an expected value is missing (order-insensitive subset check).
      * twitter_search_terms / youtube_channel_username / -company - curated
        terms and channels are valid alternatives ("both values are
        correct"); never flag as Mismatch (gap only).
      * genre                           - must not be blank for Movies / TV /
        video games; any curated non-blank set is fine; suggestions trimmed
        to the top 3.
      * primary_genre                   - valid if it is any one of the values
        in the row's own genre column.
      * released_on                     - discovery may pick a same-named
        title; a well-formed manual date wins.
      * wikipedia_page                  - page slug must match the title and
        must not conflict with the title/sub-category (e.g. an
        '(American_football)' page for an Actor is still an error).
      * facebook_page                   - /p/ and /people/ URLs are never
        valid (flagged in the manual value); 'Fanpage', profile.php and bare
        numeric-id URLs are additionally never offered as suggestions.
      * instagram_user                  - 'Fanpage' handles are ignored.
    """
    c = str(col or '').strip().lower()
    mval = _review_norm(manual_raw)
    eval_ = _review_norm(expected_raw)
    sugg = str(expected_raw if expected_raw is not None else '')

    if c == 'genre' and sugg:          # top-3 genres only in suggestions
        sugg = '\n'.join([l.strip() for l in sugg.split('\n') if l.strip()][:3])

    if not eval_ or mval == eval_:
        return True, sugg

    man_lines, exp_lines = _review_lines_ci(mval), _review_lines_ci(eval_)

    if c in ('title_sub_category', 'brand_set'):
        return exp_lines <= man_lines, sugg

    if c in ('twitter_search_terms', 'youtube_channel_username',
             'youtube_channel_company'):
        # reviewer: "both values are correct" -- curated terms/channels are
        # valid alternatives; only flag when the cell is empty (gap)
        return bool(mval), sugg

    if c == 'genre':
        return bool(mval), sugg

    if c == 'primary_genre':
        return bool(mval) and mval.strip().lower() in _review_lines_ci(manual_genre), sugg

    if c == 'released_on':
        return bool(re.match(r'^\d{4}-\d{2}-\d{2}', mval)), sugg

    if c == 'wikipedia_page':
        if not mval:
            return False, sugg
        base = title[:-6].strip() if str(title).endswith(' - DAR') else str(title)
        tslug = _review_slug(base)
        ttoks = set(re.findall(r'[a-z0-9]+', base.lower()))
        ctx = set(re.findall(r'[a-z0-9]+', (str(cat) + ' ' + str(sub_raw)).lower()))
        for ln in mval.split('\n'):
            slug = ln.rsplit('/', 1)[-1]
            if 'disambiguation' in slug.lower():
                continue
            # slug matches the title, or is a name-variant of it
            # (e.g. Kelsey_Asbille for 'Kelsey Asbille Chow')
            stoks = set(re.findall(r'[a-z0-9]+', re.sub(r'\([^)]*\)', '', slug).lower()))
            if not ((tslug and tslug in _review_slug(slug))
                    or (stoks and stoks <= ttoks)):
                continue
            par = re.search(r'\(([^)]*)\)', slug)
            if not par:
                return True, sugg
            ptoks = set(re.findall(r'[a-z0-9]+', par.group(1).lower()))
            if not ctx or (ptoks & ctx):
                return True, sugg
        return False, sugg

    if c == 'facebook_page':
        good = [l for l in sorted(exp_lines)
                if not _BAD_FB_SUGG_RE.search(l) and not _FANPAGE_RE.search(l)]
        if not mval:
            return (not good), ('\n'.join(good) if good else sugg)
        if any(_BAD_FB_MANUAL_RE.search(l) for l in man_lines):
            return False, '\n'.join(good)
        return True, sugg

    if c == 'instagram_user':
        good = [l for l in sorted(exp_lines) if not _FANPAGE_RE.search(l)]
        if not mval:
            return (not good), ('\n'.join(good) if good else sugg)
        if _FANPAGE_RE.search(mval):
            return False, '\n'.join(good)
        return True, sugg

    return False, sugg
# ==================================================================================


# categories handled by the original four review branches
_TFX_LEGACY_CATS = {'movies', 'tv shows', 'talent', 'video game'}


def _tfx_schema_for_row(r, cat):
    """Return 'beauty'/'beverages'/'sports'/'general' when the row belongs to one
    of the four NEW ingest schemas, else None (row keeps its legacy handling).
    Works even when title_category is blank, via sub_category/brand_set signals."""
    if not TFX_OK:
        return None
    c = str(cat or '').strip()
    if c.lower() in _TFX_LEGACY_CATS or 'game' in c.lower():
        return None
    key = _tfx_detect(r)
    if key:
        return key
    # category is present and belongs to the General master list
    if c and c in _TFX_MASTER:
        return 'general'
    return None


def build_review(src, auto_fetch=True, progress=None):
    """Review an uploaded manual workbook. Returns (xlsx_bytes, summary)."""
    df = _read_upload(src)
    lower_cols = {c.lower(): c for c in df.columns}
    records = df.to_dict('records')
    findings, fills = [], {}
    rows_reviewed = cells_checked = 0
    total = len(records)

    for i, r in enumerate(records):
        t = str(r.get(lower_cols.get('title', 'title'), '') or '').strip()
        if not t:
            if progress:
                progress(i + 1, total)
            continue
        rows_reviewed += 1
        cat = str(r.get(lower_cols.get('title_category', ''), '') or '')

        # ---- NEW SCHEMAS: Beauty / Beverages / Sports Teams / General ----
        tfx_key = _tfx_schema_for_row(r, cat)
        if tfx_key:
            fcat, fsub = _tfx_fill(r)
            cat_col = lower_cols.get('title_category')
            sub_col = lower_cols.get('title_sub_category')
            # backfill a MISSING category (auto-understood from the row) as a Gap
            if cat_col is not None:
                cells_checked += 1
                if not str(r.get(cat_col) or '').strip() and fcat:
                    findings.append(dict(row=i + 2, title=t, column=cat_col,
                                         status='Gap', current='', suggested=fcat))
                    fills[(i, cat_col)] = 'Gap'
                    r[cat_col] = fcat   # validate the rest against the filled value
            # backfill a MISSING sub-category the same way
            if sub_col is not None:
                cells_checked += 1
                if not str(r.get(sub_col) or '').strip() and fsub:
                    findings.append(dict(row=i + 2, title=t, column=sub_col,
                                         status='Gap', current='', suggested=fsub))
                    fills[(i, sub_col)] = 'Gap'
                    r[sub_col] = fsub
            # dropdown / template-logic validation
            try:
                tfx_findings = _tfx_validate(r, tfx_key, _TFX_RULES)
            except Exception as _e:
                logging.warning(f"tfx validate failed on row {i + 2}: {_e}")
                tfx_findings = []
            n_rules = len(_TFX_RULES.get('schemas', {}).get(tfx_key, {}).get('rules', []))
            cells_checked += max(n_rules - 2, 0)  # category+sub already counted
            for fd in tfx_findings:
                src_col = lower_cols.get(str(fd.get('field', '')).lower())
                if not src_col:
                    continue
                if fills.get((i, src_col)):
                    continue  # already flagged by the backfill above
                status = 'Gap' if fd.get('status') == 'gap' else 'Mismatch'
                findings.append(dict(
                    row=i + 2, title=t, column=src_col, status=status,
                    current=str(r.get(src_col) if r.get(src_col) is not None else ''),
                    suggested=str(fd.get('expected') or '')))
                fills[(i, src_col)] = status
            if progress:
                progress(i + 1, total)
            continue
        # -------------------------------------------------------------------

        is_talent = 'talent' in cat.lower()
        is_game = 'game' in cat.lower()
        is_movie = (not is_talent) and (not is_game) and 'tv' not in cat.lower()
        sub = _sub_parts(r.get(lower_cols.get('title_sub_category', '')))

        if is_game:
            meta = dict(fetch_game(t) or {}) if auto_fetch else {}
            # manual sub lines fill discovery gaps (developer / platforms)
            if not meta.get('developer') and sub.get('Developer'):
                meta['developer'] = sub['Developer']
            if not meta.get('platforms'):
                plats = [l.split(' - ', 1)[1] for l in
                         str(r.get(lower_cols.get('title_sub_category', ''), '') or '').split('\n')
                         if l.startswith('Platform - ')]
                if plats:
                    meta['platforms'] = plats
            if not meta.get('network'):
                meta['network'] = str(r.get(lower_cols.get('network', ''), '') or '').strip()
            g = str(r.get(lower_cols.get('genre', ''), '') or '').strip()
            if not meta.get('genre') and g:
                meta['genre'] = g
            expected = make_row(t, False, '', meta, game=True)
            for col in GAME_COLUMNS:
                if col in REVIEW_SKIP_COLS or col.lower() not in lower_cols:
                    continue
                src_col = lower_cols[col.lower()]
                mval = _review_norm(r.get(src_col))
                cells_checked += 1
                ok, sugg = _review_compare(
                    col, r.get(src_col), expected.get(col), title=t, cat=cat,
                    manual_genre=r.get(lower_cols.get('genre', ''), ''),
                    sub_raw=r.get(lower_cols.get('title_sub_category', ''), ''))
                if ok:
                    continue
                status = 'Gap' if not mval else 'Mismatch'
                findings.append(dict(
                    row=i + 2, title=t, column=src_col, status=status,
                    current=str(r.get(src_col) if r.get(src_col) is not None else ''),
                    suggested=sugg))
                fills[(i, src_col)] = status
            if progress:
                progress(i + 1, total)
            continue

        if is_talent:
            meta = dict(fetch_person(t) or {}) if auto_fetch else {}
            # manual sub-category lines fill classification gaps
            if not meta.get('gender') and sub.get('Gender'):
                meta['gender'] = 'Gender - ' + sub['Gender']
            if sub.get('Talent Type') or sub.get('Talent Subtype'):
                lines = [x for x in (
                    ('Talent Subtype - ' + sub['Talent Subtype']) if sub.get('Talent Subtype') else '',
                    ('Gender - ' + sub['Gender']) if sub.get('Gender') else '',
                    ('Talent Type - ' + sub['Talent Type']) if sub.get('Talent Type') else '') if x]
                if not (meta.get('occupations') or meta.get('sports')):
                    meta['title_sub_category'] = '\n'.join(lines)
            expected = make_row(t, False, '', meta, talent=True)
            for col in TALENT_COLUMNS:
                if col in REVIEW_SKIP_COLS or col.lower() not in lower_cols:
                    continue
                src_col = lower_cols[col.lower()]
                mval = _review_norm(r.get(src_col))
                cells_checked += 1
                ok, sugg = _review_compare(
                    col, r.get(src_col), expected.get(col), title=t, cat=cat,
                    manual_genre=r.get(lower_cols.get('genre', ''), ''),
                    sub_raw=r.get(lower_cols.get('title_sub_category', ''), ''))
                if ok:
                    continue
                status = 'Gap' if not mval else 'Mismatch'
                findings.append(dict(
                    row=i + 2, title=t, column=src_col, status=status,
                    current=str(r.get(src_col) if r.get(src_col) is not None else ''),
                    suggested=sugg))
                fills[(i, src_col)] = status
            if progress:
                progress(i + 1, total)
            continue

        # soft hints from the manual row: fill discovery gaps, never override
        hints = {}
        if is_movie:
            if sub.get('Release'):
                hints['release_scale'] = sub['Release']
        else:
            if sub.get('Program Type'):
                hints['program_type'] = sub['Program Type']
        if sub.get('Language Type'):
            hints['original_language'] = 'en' if sub['Language Type'] == 'English' else 'xx'
        rel = str(r.get(lower_cols.get('released_on', ''), '') or '').strip()
        if rel and rel.lower() != 'nan':
            hints['released_on'] = rel[:10]

        meta = {}
        if auto_fetch:
            tt = re.search(r'tt\d{5,}', str(r.get(lower_cols.get('imdb_id', ''), '') or ''))
            if tt:
                meta = dict(fetch_metadata_by_tt(tt.group(0), is_movie, t) or {})
            else:
                meta = dict(fetch_metadata(t, is_movie) or {})
        for k, v in hints.items():
            if meta.get(k) in (None, ''):
                meta[k] = v

        exp_net = str(meta.get('network') or
                      r.get(lower_cols.get('network', ''), '') or '').strip()
        if exp_net and not meta.get('network'):
            meta['network'] = exp_net
        expected = make_row(t, is_movie, exp_net, meta)

        for col in (COLUMNS if is_movie else TV_COLUMNS):
            if col in REVIEW_SKIP_COLS or col.lower() not in lower_cols:
                continue
            src_col = lower_cols[col.lower()]
            mval = _review_norm(r.get(src_col))
            cells_checked += 1
            ok, sugg = _review_compare(
                col, r.get(src_col), expected.get(col), title=t, cat=cat,
                manual_genre=r.get(lower_cols.get('genre', ''), ''),
                sub_raw=r.get(lower_cols.get('title_sub_category', ''), ''))
            if ok:
                continue
            status = 'Gap' if not mval else 'Mismatch'
            findings.append(dict(
                row=i + 2, title=t, column=src_col, status=status,
                current=str(r.get(src_col) if r.get(src_col) is not None else ''),
                suggested=sugg))
            fills[(i, src_col)] = status
        if progress:
            progress(i + 1, total)

    # ---------------- build the output workbook ----------------
    import openpyxl as _oxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    RED = PatternFill('solid', start_color='FFFFC7CE')      # mismatch
    AMBER = PatternFill('solid', start_color='FFFFEB9C')    # gap
    HDR = PatternFill('solid', start_color='FF1F2A44')
    HDR_FONT = Font(color='FFFFFFFF', bold=True)

    wb = _oxl.Workbook()

    # Summary
    ws = wb.active
    ws.title = 'Summary'
    gaps = sum(1 for f in findings if f['status'] == 'Gap')
    mism = len(findings) - gaps
    ws.append(['Manual File Review — Findings Summary'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    for k, v in [('Reviewed at', datetime.now().strftime('%Y-%m-%d %H:%M')),
                 ('Auto-discovery', 'ON' if auto_fetch else 'OFF'),
                 ('Rows reviewed', rows_reviewed),
                 ('Cells checked', cells_checked),
                 ('Cells OK', cells_checked - len(findings)),
                 ('Gaps (empty, value suggested)', gaps),
                 ('Mismatches (differs from expected)', mism)]:
        ws.append([k, v])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append([])
    ws.append(['Legend'])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append(['Amber cell', 'Gap — the tool found a value your file is missing'])
    ws.cell(ws.max_row, 1).fill = AMBER
    ws.append(['Red cell', 'Mismatch — differs from template/discovered value (see Findings)'])
    ws.cell(ws.max_row, 1).fill = RED
    from collections import Counter as _Counter
    by_col = _Counter(f['column'] for f in findings)
    if by_col:
        ws.append([])
        ws.append(['Findings by column'])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        for col, n in by_col.most_common():
            ws.append([col, n])
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 64

    # Reviewed copy with highlights
    ws2 = wb.create_sheet('Reviewed')
    cols = list(df.columns)
    ws2.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws2.cell(1, c)
        cell.fill, cell.font = HDR, HDR_FONT
    for i, r in enumerate(records):
        ws2.append(['' if (r.get(c) is None or str(r.get(c)) == 'nan') else r.get(c)
                    for c in cols])
        for j, c in enumerate(cols, start=1):
            st = fills.get((i, c))
            if st:
                ws2.cell(i + 2, j).fill = RED if st == 'Mismatch' else AMBER
    ws2.freeze_panes = 'A2'

    # Findings detail
    ws3 = wb.create_sheet('Findings')
    ws3.append(['Row', 'Title', 'Column', 'Type', 'Current Value', 'Suggested Value'])
    for c in range(1, 7):
        cell = ws3.cell(1, c)
        cell.fill, cell.font = HDR, HDR_FONT
    for f in findings:
        ws3.append([f['row'], f['title'], f['column'], f['status'],
                    f['current'], f['suggested']])
        ws3.cell(ws3.max_row, 4).fill = RED if f['status'] == 'Mismatch' else AMBER
        for c in (5, 6):
            ws3.cell(ws3.max_row, c).alignment = Alignment(wrap_text=True, vertical='top')
    widths = [6, 34, 26, 11, 60, 60]
    for c, w in enumerate(widths, start=1):
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{max(ws3.max_row, 1)}"

    out = BytesIO()
    wb.save(out)
    summary = {'rows': rows_reviewed, 'cells_checked': cells_checked,
               'gaps': gaps, 'mismatches': mism, 'ok': cells_checked - len(findings)}
    return out.getvalue(), summary


@app.route('/review')
def review_page():
    warm_upcoming()
    return render_template('review.html')


@app.route('/api/review_async', methods=['POST'])
def review_async():
    """Kick off a manual-file review in the background; returns a job id."""
    _prune_jobs()
    if not request.files.get('file'):
        return jsonify({'error': 'Please upload the manually prepared .xlsx or .csv file.'}), 400
    f = request.files['file']
    payload = {
        'bytes': f.read(), 'filename': f.filename,
        'auto_fetch': request.form.get('autoFetch', 'true').lower() != 'false',
    }
    jid = uuid.uuid4().hex[:12]
    with _JOBS_LOCK:
        _JOBS[jid] = {'status': 'running', 'done': 0, 'total': 0, 'error': None,
                      'file': None, 'filename': None, 'rows': None,
                      'summary': None, 'created': time.time()}
    threading.Thread(target=_run_generation, args=(jid, 'review', payload),
                     daemon=True).start()
    return jsonify({'job_id': jid})


# ============================ background jobs =============================
# In-memory job store for full-file generation. Runs in a daemon thread so long
# auto-discovery runs don't hit the request timeout. IMPORTANT: run gunicorn with
# a SINGLE worker + threads so this store is shared, e.g.:
#   web: gunicorn app:app --workers 1 --threads 8 --timeout 120
_JOBS = {}
_JOBS_LOCK = threading.Lock()
_JOB_TTL = 1800  # seconds to keep a finished job's file in memory


def _job_set(jid, **kw):
    with _JOBS_LOCK:
        if jid in _JOBS:
            _JOBS[jid].update(kw)


def _prune_jobs():
    now = time.time()
    with _JOBS_LOCK:
        for k in [k for k, v in _JOBS.items() if now - v.get('created', now) > _JOB_TTL]:
            _JOBS.pop(k, None)


def _run_generation(jid, kind, payload):
    try:
        def prog(done, total):
            _job_set(jid, done=done, total=total)

        if kind == 'review':
            data, summary = build_review((payload['bytes'], payload['filename']),
                                         payload['auto_fetch'], progress=prog)
            _job_set(jid, status='done', file=data, rows=summary['rows'],
                     summary=summary,
                     filename=f"Reviewed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            return
        max_titles = PREVIEW_MAX_TITLES if payload.get('preview') else None
        if kind == 'file':
            rows = build_rows_from_upload(
                (payload['bytes'], payload['filename']),
                payload['include_dar'], payload['auto_fetch'], progress=prog,
                max_titles=max_titles,
                default_kind=payload.get('title_type', 'movie'))
        else:
            rows = build_rows_from_titles(payload['data'], max_titles=max_titles,
                                          progress=prog)

        if not rows:
            _job_set(jid, status='error', error='No titles provided')
            return
        if payload.get('preview'):
            _job_set(jid, status='done',
                     preview=_preview_payload(rows, payload.get('preview_limited', False)),
                     rows=len(rows))
            return
        # how many rows actually got social/discovery data -- surfaces
        # rate-limit problems instead of silently exporting blank socials
        _soc = ('twitter_handle', 'instagram_user', 'facebook_page',
                'youtube_channel_username', 'wikipedia_page', 'tiktok_user')
        enriched = sum(1 for r in rows
                       if any(str(r.get(c) or '').strip() for c in _soc))
        out = _rows_to_workbook(rows)
        _job_set(jid, status='done', file=out.getvalue(), rows=len(rows),
                 enriched=enriched,
                 filename=f"Titles_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    except Exception as e:  # noqa: BLE001
        logging.error(f"generation job {jid} failed: {e}")
        _job_set(jid, status='error', error=str(e))


@app.route('/api/generate_async', methods=['POST'])
def generate_async():
    """Kick off full-file generation in the background; returns a job id."""
    _prune_jobs()
    jid = uuid.uuid4().hex[:12]
    if request.files.get('file'):
        f = request.files['file']
        payload = {
            'bytes': f.read(), 'filename': f.filename,
            'include_dar': request.form.get('includeDar', 'true').lower() != 'false',
            'auto_fetch': request.form.get('autoFetch', 'false').lower() == 'true',
            'title_type': _norm_kind(request.form.get('titleType')),
        }
        kind = 'file'
    else:
        payload = {'data': request.get_json(silent=True) or {}}
        kind = 'titles'
    with _JOBS_LOCK:
        _JOBS[jid] = {'status': 'running', 'done': 0, 'total': 0, 'error': None,
                      'file': None, 'filename': None, 'rows': None, 'created': time.time()}
    threading.Thread(target=_run_generation, args=(jid, kind, payload), daemon=True).start()
    return jsonify({'job_id': jid})


@app.route('/api/preview_async', methods=['POST'])
def preview_async():
    """Kick off a preview (first PREVIEW_MAX_TITLES titles, incl. auto-discover
    enrichment) in the background; returns a job id. Poll /api/job/<jid> --
    when done the job carries a 'preview' payload. This keeps long
    auto-discovery lookups out of the HTTP request, which the platform
    kills after ~30s (the old 502 'server timed out' error)."""
    _prune_jobs()
    jid = uuid.uuid4().hex[:12]
    if request.files.get('file'):
        f = request.files['file']
        payload = {
            'bytes': f.read(), 'filename': f.filename,
            'include_dar': request.form.get('includeDar', 'true').lower() != 'false',
            'auto_fetch': request.form.get('autoFetch', 'false').lower() == 'true',
            'title_type': _norm_kind(request.form.get('titleType')),
            'preview': True, 'preview_limited': True,
        }
        kind = 'file'
    else:
        data = request.get_json(silent=True) or {}
        n_src = len([t for t in data.get('titles', []) if t and t.strip()])
        payload = {'data': data, 'preview': True,
                   'preview_limited': n_src > PREVIEW_MAX_TITLES}
        kind = 'titles'
    with _JOBS_LOCK:
        _JOBS[jid] = {'status': 'running', 'done': 0, 'total': 0, 'error': None,
                      'file': None, 'filename': None, 'rows': None, 'created': time.time()}
    threading.Thread(target=_run_generation, args=(jid, kind, payload), daemon=True).start()
    return jsonify({'job_id': jid})


@app.route('/api/job/<jid>')
def job_status(jid):
    with _JOBS_LOCK:
        j = _JOBS.get(jid)
        if not j:
            return jsonify({'error': 'Unknown or expired job'}), 404
        eta = None
        if j['status'] == 'running' and j.get('done') and j.get('total'):
            elapsed = time.time() - j.get('created', time.time())
            eta = max(0, int(elapsed / j['done'] * (j['total'] - j['done'])))
        return jsonify({'status': j['status'], 'done': j['done'], 'total': j['total'],
                        'error': j['error'], 'rows': j.get('rows'),
                        'summary': j.get('summary'), 'eta_seconds': eta,
                        'preview': j.get('preview'),
                        'enriched': j.get('enriched')})


@app.route('/api/job/<jid>/download')
def job_download(jid):
    with _JOBS_LOCK:
        j = _JOBS.get(jid)
        if not j or j['status'] != 'done' or not j['file']:
            return jsonify({'error': 'File not ready'}), 404
        data, fn = j['file'], j['filename']
    return send_file(BytesIO(data),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fn)


if __name__ == '__main__':
    app.run(debug=True)
