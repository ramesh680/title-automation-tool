"""
reference_data.py
-----------------
Loads the Ops ingest templates as AUTHORITATIVE reference data:

  reference/film_ingest_template.xlsx  ("NEW Ingest Template - Film w/ Language")
    - DropDown sheet: 271 studios -> Studio Type, YouTube (Company), parent
      Company, Twitter Keyword String, Reddit Keyword String
    - 'Media Conglomerate Brand Sets' sheet: per-studio DAR roll-up lines

  reference/tv_ingest_template.xlsx    ("New Brand Definitions Ingest Template - TV")
    - DropdownLists sheet: 228 networks -> Network Type, Ticker, YouTube,
      Twitter/Reddit Keyword Strings; primary-genre Order of Operations
    - 'Conglomerate Brand Sets' sheet: per-network DAR roll-up block

To update the tool's logic, Ops replaces these files in the repo's
reference/ folder -- no code change needed. Paths can be overridden with the
REFERENCE_DIR / FILM_TEMPLATE / TV_TEMPLATE environment variables.

Everything fails soft: if the files are missing or unreadable, the app falls
back to its built-in tables.
"""

import logging
import os

log = logging.getLogger(__name__)

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

REF_DIR = os.getenv("REFERENCE_DIR",
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), "reference"))
FILM_TEMPLATE = os.getenv("FILM_TEMPLATE", os.path.join(REF_DIR, "film_ingest_template.xlsx"))
TV_TEMPLATE = os.getenv("TV_TEMPLATE", os.path.join(REF_DIR, "tv_ingest_template.xlsx"))
TALENT_TEMPLATE = os.getenv("TALENT_TEMPLATE", os.path.join(REF_DIR, "talent_ingest_template.xlsx"))
GAME_TEMPLATE = os.getenv("GAME_TEMPLATE", os.path.join(REF_DIR, "videogame_ingest_template.xlsx"))

FILM_STUDIOS = {}       # studio(lower) -> dict(studio_type, youtube, company, twitter_clause, reddit_clause)
FILM_ROLLUPS = {}       # studio(lower) -> "Roll-Up 1\nRoll-Up 2[\nRoll-Up 3]"
TV_NETWORKS = {}        # network(lower) -> dict(network_type, ticker, youtube, twitter_clause, reddit_clause)
TV_CONGLOMERATES = {}   # network(lower) -> multi-line brand_set block
TV_PRIMARY_ORDER = []   # [(genre, mapped_primary), ...] in Order of Operations
TALENT_TYPES = []       # ['Talent Type - Actor', ...]
TALENT_SUBTYPES = []    # ['Talent Subtype - Athlete - Basketball', ...]
TALENT_BRAND_SETS = []  # optional extra talent brand sets
GAME_DEVELOPERS = {}    # dev(lower, no prefix) -> dict(twitter_clause, reddit_clause, youtube)
GAME_PUBLISHERS = {}    # publisher(lower) -> dict(youtube, twitter_clause)
GAME_GENRE_MAP = {}     # genre(lower) -> Primary Genre
GAME_PLATFORMS = []     # ['Platform - PC', ...]

LOADED = False


def _s(v):
    return str(v).strip() if v is not None else ""


def _load_film():
    wb = openpyxl.load_workbook(FILM_TEMPLATE, data_only=True)
    ws = wb["DropDown"]
    # main studio table: J=Studio, K=Studio Type, L=YouTube, M=Company, N=Twitter
    for r in range(2, ws.max_row + 1):
        s = _s(ws.cell(r, 10).value)
        if not s:
            continue
        FILM_STUDIOS[s.lower()] = dict(
            name=s,
            studio_type=_s(ws.cell(r, 11).value),
            youtube=_s(ws.cell(r, 12).value),
            company=_s(ws.cell(r, 13).value),
            twitter_clause=_s(ws.cell(r, 14).value),
            reddit_clause="",
        )
    # parallel reddit table: R=Studio ... V=Reddit Keyword String
    for r in range(2, ws.max_row + 1):
        s = _s(ws.cell(r, 18).value)
        if s and s.lower() in FILM_STUDIOS:
            FILM_STUDIOS[s.lower()]["reddit_clause"] = _s(ws.cell(r, 22).value)
    # per-studio conglomerate roll-ups (column-oriented)
    ws2 = wb["Media Conglomerate Brand Sets"]
    for c in range(2, ws2.max_column + 1):
        studio = _s(ws2.cell(2, c).value)
        if not studio:
            continue
        lines = [_s(ws2.cell(rr, c).value) for rr in (3, 4, 5)]
        lines = [l for l in lines if l]
        if lines:
            FILM_ROLLUPS[studio.lower()] = "\n".join(lines)


def _load_tv():
    wb = openpyxl.load_workbook(TV_TEMPLATE, data_only=True)
    ws = wb["DropdownLists"]
    # B=Companies, C=Network, D=Network Type, E=Ticker, F=YouTube, G=Twitter
    for r in range(2, ws.max_row + 1):
        n = _s(ws.cell(r, 3).value)
        if not n:
            continue
        TV_NETWORKS[n.lower()] = dict(
            name=n,
            company=_s(ws.cell(r, 2).value),
            network_type=_s(ws.cell(r, 4).value),
            ticker=_s(ws.cell(r, 5).value),
            youtube=_s(ws.cell(r, 6).value),
            twitter_clause=_s(ws.cell(r, 7).value),
            reddit_clause="",
        )
    # parallel reddit table: R=Network ... V=Reddit Keyword String
    for r in range(2, ws.max_row + 1):
        n = _s(ws.cell(r, 18).value)
        if n and n.lower() in TV_NETWORKS:
            TV_NETWORKS[n.lower()]["reddit_clause"] = _s(ws.cell(r, 22).value)
    # primary-genre Order of Operations: N=genre, O=Mapped Primary Genre
    for r in range(2, ws.max_row + 1):
        g, m = _s(ws.cell(r, 14).value), _s(ws.cell(r, 15).value)
        if g and m:
            TV_PRIMARY_ORDER.append((g, m))
    # per-network conglomerate brand-set block
    ws2 = wb["Conglomerate Brand Sets"]
    for r in range(2, ws2.max_row + 1):
        n, b = _s(ws2.cell(r, 1).value), _s(ws2.cell(r, 2).value)
        if not n or b in ("", "None"):
            continue
        for key in (n.lower(), n.split("\n")[0].strip().lower()):
            TV_CONGLOMERATES.setdefault(key, b)


def _load_talent():
    wb = openpyxl.load_workbook(TALENT_TEMPLATE, data_only=True)
    ws = wb["DropDown"]
    for r in range(2, ws.max_row + 1):
        sub = _s(ws.cell(r, 6).value)     # 'Talent Subtype - ...'
        typ = _s(ws.cell(r, 7).value)     # 'Talent Type - ...'
        bs = _s(ws.cell(r, 4).value)      # brand sets
        if sub and sub not in TALENT_SUBTYPES:
            TALENT_SUBTYPES.append(sub)
        if typ and typ not in TALENT_TYPES:
            TALENT_TYPES.append(typ)
        if bs and bs not in TALENT_BRAND_SETS:
            TALENT_BRAND_SETS.append(bs)


def _dev_key(v):
    return _s(v).replace("Developer - ", "").strip().lower()


def _load_game():
    wb = openpyxl.load_workbook(GAME_TEMPLATE, data_only=True)
    ws = wb["DropDown"]
    for r in range(2, ws.max_row + 1):
        dev = _s(ws.cell(r, 2).value)      # 'Developer - X'
        if dev:
            GAME_DEVELOPERS.setdefault(_dev_key(dev), {})["twitter_clause"] = \
                _s(ws.cell(r, 3).value)
        dev_r = _s(ws.cell(r, 11).value)   # reddit table
        if dev_r:
            GAME_DEVELOPERS.setdefault(_dev_key(dev_r), {})["reddit_clause"] = \
                _s(ws.cell(r, 12).value)
        dev_y = _s(ws.cell(r, 14).value)   # dev youtube table
        if dev_y:
            GAME_DEVELOPERS.setdefault(_dev_key(dev_y), {})["youtube"] = \
                _s(ws.cell(r, 15).value)
        pub = _s(ws.cell(r, 4).value)
        if pub:
            GAME_PUBLISHERS[pub.lower()] = dict(
                youtube=_s(ws.cell(r, 5).value),
                twitter_clause=_s(ws.cell(r, 6).value))
        plat = _s(ws.cell(r, 8).value)
        if plat and plat not in GAME_PLATFORMS:
            GAME_PLATFORMS.append(plat)
        g, p = _s(ws.cell(r, 9).value), _s(ws.cell(r, 10).value)
        if g and p:
            GAME_GENRE_MAP[g.lower()] = p


def load():
    """(Re)load the templates. Returns True when any reference data loaded."""
    global LOADED
    if openpyxl is None:
        return False
    try:
        if os.path.exists(FILM_TEMPLATE):
            _load_film()
        else:
            log.warning("film template not found: %s", FILM_TEMPLATE)
        if os.path.exists(TV_TEMPLATE):
            _load_tv()
        else:
            log.warning("tv template not found: %s", TV_TEMPLATE)
        if os.path.exists(TALENT_TEMPLATE):
            _load_talent()
        else:
            log.warning("talent template not found: %s", TALENT_TEMPLATE)
        if os.path.exists(GAME_TEMPLATE):
            _load_game()
        else:
            log.warning("game template not found: %s", GAME_TEMPLATE)
        LOADED = bool(FILM_STUDIOS or TV_NETWORKS)
        log.info("reference templates loaded: %d studios, %d networks, "
                 "%d film roll-ups, %d tv conglomerates, %d genre mappings",
                 len(FILM_STUDIOS), len(TV_NETWORKS), len(FILM_ROLLUPS),
                 len(TV_CONGLOMERATES), len(TV_PRIMARY_ORDER))
    except Exception as e:  # noqa: BLE001
        log.warning("reference template load failed: %s", e)
    return LOADED


# ---------------- lookups ----------------
def film_studio(studio):
    return FILM_STUDIOS.get(_s(studio).lower())


def film_rollup(studio):
    return FILM_ROLLUPS.get(_s(studio).lower(), "")


def tv_network(network):
    return TV_NETWORKS.get(_s(network).lower())


def tv_conglomerate(network):
    return TV_CONGLOMERATES.get(_s(network).lower(), "")


def tv_primary_genre(genres):
    """First genre in the Order-of-Operations list that the title has wins;
    its Mapped Primary Genre is the answer ('' when nothing matches)."""
    gs = {_s(g) for g in genres}
    for g, mapped in TV_PRIMARY_ORDER:
        if g in gs and mapped != "N/A":
            return mapped
    return ""


def game_developer(name):
    return GAME_DEVELOPERS.get(_dev_key(name))


def game_publisher(name):
    return GAME_PUBLISHERS.get(_s(name).lower())


def game_primary_genre(genre):
    return GAME_GENRE_MAP.get(_s(genre).lower(), "")


def game_platform_for(term):
    """Match a discovered platform label to the template's 'Platform - X'.
    Exact match wins before fuzzy (so 'Switch 2' doesn't hit 'Switch')."""
    t = _s(term).lower()
    if not t:
        return ""
    for p in GAME_PLATFORMS:
        if p.split(" - ", 1)[1].lower() == t:
            return p
    for p in GAME_PLATFORMS:
        tail = p.split(" - ", 1)[1].lower()
        if tail in t or t in tail:
            return p
    return ""


def talent_subtype_for(kind, term):
    """Find the canonical 'Talent Subtype - <kind> - <x>' entry matching a
    discovered term (e.g. kind='Athlete', term='Basketball')."""
    t = _s(term).lower()
    if not t:
        return ""
    prefix = f"talent subtype - {_s(kind).lower()} - "
    for sub in TALENT_SUBTYPES:
        sl = sub.lower()
        if sl.startswith(prefix):
            tail = sl[len(prefix):]
            if tail == t or t in tail or tail in t:
                return sub
    return ""


load()
