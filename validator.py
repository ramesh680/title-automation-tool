"""
validator.py
------------
Workbook Sheet Validator for Title Data exports.

Upload an .xlsx/.csv, apply a rules-JSON, get back the same workbook with
failing cells highlighted plus a "Validation Summary" sheet -- mirroring the
media-tools-hub "Official Profile Finder / Data Ops Larger Project Validation"
tool, tailored to this app's 42-column schema.

The rules engine is data-driven: each rule is {sheet, column, check, ...params}.
Deterministic checks run fully offline. "lookup_*" checks that would need
external data (IMDb datasets, Metacritic, live Wikidata, YouTube API) currently
validate presence + format and are flagged as warnings; the real lookups plug
in later (they can reuse metadata_fetcher / the repo logic).
"""

import csv
import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ----- severities & highlight colours -------------------------------------
SEV_FAIL = "fail"     # hard failure -> red
SEV_WARN = "warn"     # needs review / pending lookup -> amber

FILL_FAIL = PatternFill("solid", fgColor="FFF4C7C3")   # soft red
FILL_WARN = PatternFill("solid", fgColor="FFFFE8A3")   # soft amber
FILL_HEAD = PatternFill("solid", fgColor="FF7C5CFF")   # brand violet

APPROVED_CATEGORIES = {"movies", "tv shows"}
# extend with the master Title Category list from the General ingest template
# (Health & Beauty, Beverages, Sports Franchise, Talent, Video Game, + 44 more)
try:
    from titleforge_ingest_ext import ALL_TITLE_CATEGORIES as _TFX_ALL
    APPROVED_CATEGORIES |= {c.lower() for c in _TFX_ALL}
except Exception:  # fail soft: keep the original two if the module is absent
    pass

# Default rules, tailored to THIS app's export schema (column names differ from
# the upstream tool: instagram_user / twitter_handle / tiktok_user / threads_page,
# metacritic, wikipedia_page). Same check semantics.
DEFAULT_RULES = {
    "rules": [
        {"sheet": "*", "column": "title", "check": "not_blank_and_not_placeholder",
         "tokens": ["#NA", "N/A"], "message": "Title cannot be blank, #NA, or N/A."},
        {"sheet": "*", "column": "title_category", "check": "approved_category",
         "message": "title_category must be present and one of the approved categories "
                    "(Movies, TV Shows, Talent, Video Game, Health & Beauty, Beverages, "
                    "Sports Franchise, or the General master list)."},
        {"sheet": "*", "column": "brand_set", "check": "dar_or_competitive_brand_set",
         "message": "DAR titles need 'Pristine DAR Brands'; other titles need 'Competitive View'."},
        {"sheet": "*", "column": "companies", "check": "dar_company_rule",
         "message": "DAR titles must have companies = 'Pristine Brand'."},
        {"sheet": "*", "column": "imdb_id", "alternate_column": "imdb_url",
         "check": "imdb_ttcode_format", "applies_to": ["Movies", "TV Shows"],
         "message": "IMDb value should be an IMDb title URL / ttNNNNNNN code."},
        {"sheet": "*", "column": "metacritic", "check": "metacritic_url_format",
         "message": "metacritic value should be a metacritic.com movie/tv URL."},
        {"sheet": "*", "column": "wikipedia_page",
         "check": "english_wikipedia_url_matches_title", "accepted_host": "en.wikipedia.org",
         "message": "Wikipedia URLs must be en.wikipedia.org/wiki/... and match the title."},
        {"sheet": "*", "column": "url_managers",
         "check": "contains_companies_and_platform_accounts",
         "company_column": "companies",
         "exclude_company_values": ["Unknown", "Pristine Brand"],
         "platform_columns": ["facebook_page", "youtube_channel_company",
                              "instagram_user", "twitter_handle", "tiktok_user", "threads_page"],
         "message": "url_managers must reference the row's social accounts (Unknown / Pristine Brand skipped)."},
    ]
}


# ----- helpers -------------------------------------------------------------
def _s(v):
    return "" if v is None else str(v).strip()


def _norm(v):
    return _s(v).lower()


def _is_dar(row):
    return " - dar" in _norm(row.get("title"))


def _row_get(row, col):
    """Case-insensitive row lookup."""
    if col in row:
        return row[col]
    low = {k.lower(): k for k in row}
    return row.get(low.get((col or "").lower(), ""), "")


# ----- individual checks ---------------------------------------------------
# each returns (severity_or_None, message)  -> None severity means "pass"
def _chk_not_blank_and_not_placeholder(val, row, rule):
    tokens = [t.upper() for t in rule.get("tokens", [])]
    if _s(val) == "" or _s(val).upper() in tokens:
        return SEV_FAIL, rule.get("message", "Value is blank or a placeholder.")
    return None, ""


def _chk_approved_category(val, row, rule):
    approved = set(a.lower() for a in rule.get("approved", APPROVED_CATEGORIES))
    if _norm(val) not in approved:
        return SEV_FAIL, rule.get("message", "Category not approved.")
    return None, ""


def _chk_dar_or_competitive_brand_set(val, row, rule):
    v = _norm(val)
    if _is_dar(row):
        if "pristine dar brands" not in v:
            return SEV_FAIL, rule.get("message")
    else:
        if "competitive view" not in v:
            return SEV_FAIL, rule.get("message")
    return None, ""


def _chk_dar_company_rule(val, row, rule):
    if _is_dar(row) and _norm(val) != "pristine brand":
        return SEV_FAIL, rule.get("message", "DAR company must be 'Pristine Brand'.")
    return None, ""


_TT_RE = re.compile(r"tt\d{5,}", re.I)


def _chk_imdb_ttcode_format(val, row, rule):
    applies = [a.lower() for a in rule.get("applies_to", [])]
    if applies and _norm(_row_get(row, "title_category")) not in applies:
        return None, ""
    v = _s(val) or _s(_row_get(row, rule.get("alternate_column", "")))
    if v == "":
        return SEV_WARN, "IMDb id missing (lookup from title pending)."
    if not _TT_RE.search(v):
        return SEV_FAIL, rule.get("message", "IMDb value malformed.")
    return None, ""


_MC_RE = re.compile(r"metacritic\.com/(movie|tv)/", re.I)


def _chk_metacritic_url_format(val, row, rule):
    v = _s(val)
    if v == "":
        return SEV_WARN, "Metacritic URL missing (lookup from title pending)."
    if not _MC_RE.search(v):
        return SEV_FAIL, rule.get("message", "Metacritic URL malformed.")
    return None, ""


def _slug(text):
    return re.sub(r"[^a-z0-9]+", "", _norm(text))


def _chk_english_wikipedia_url_matches_title(val, row, rule):
    v = _s(val)
    host = rule.get("accepted_host", "en.wikipedia.org")
    if v == "":
        return SEV_WARN, "Wikipedia URL missing (lookup from title pending)."
    if f"{host}/wiki/" not in v.lower():
        return SEV_FAIL, rule.get("message", "Not an English Wikipedia article URL.")
    # soft title-match: compare article slug to title
    article = v.rsplit("/wiki/", 1)[-1].split("#")[0]
    art = _slug(article.replace("_", " ").split("(")[0])
    title = _slug(re.sub(r"\s*-\s*dar\s*$", "", _s(_row_get(row, "title")), flags=re.I))
    if art and title and art != title and title not in art and art not in title:
        return SEV_WARN, "Wikipedia article name does not obviously match the title."
    return None, ""


def _chk_contains_companies_and_platform_accounts(val, row, rule):
    companies = _s(_row_get(row, rule.get("company_column", "companies")))
    excluded = [e.lower() for e in rule.get("exclude_company_values", [])]
    if companies.lower() in excluded or companies == "":
        return None, ""  # skipped by rule
    present = []
    for col in rule.get("platform_columns", []):
        pv = _s(_row_get(row, col))
        if pv:
            # use the most identifying token (handle or last URL path segment)
            token = pv.split("|")[0].strip().rstrip("/").rsplit("/", 1)[-1]
            present.append((col, token))
    if not present:
        return None, ""  # nothing to reference
    um = _norm(val)
    if um == "":
        return SEV_FAIL, "url_managers is blank but social accounts exist for a real company."
    missing = [c for c, tok in present if tok and _norm(tok) not in um]
    if missing:
        return SEV_WARN, "url_managers does not reference: " + ", ".join(missing)
    return None, ""


CHECKS = {
    "not_blank_and_not_placeholder": _chk_not_blank_and_not_placeholder,
    "approved_category": _chk_approved_category,
    "dar_or_competitive_brand_set": _chk_dar_or_competitive_brand_set,
    "dar_company_rule": _chk_dar_company_rule,
    "imdb_ttcode_format": _chk_imdb_ttcode_format,
    "lookup_imdb_ttcode_from_title": _chk_imdb_ttcode_format,      # alias
    "metacritic_url_format": _chk_metacritic_url_format,
    "lookup_metacritic_url_from_title": _chk_metacritic_url_format,  # alias
    "english_wikipedia_url_matches_title": _chk_english_wikipedia_url_matches_title,
    "wikidata_english_wikipedia_url_matches_title": _chk_english_wikipedia_url_matches_title,
    "contains_companies_and_platform_accounts": _chk_contains_companies_and_platform_accounts,
}


# ----- workbook loading ----------------------------------------------------
def _load_workbook(file_storage):
    """Return an openpyxl Workbook from an uploaded .xlsx or .csv."""
    name = (getattr(file_storage, "filename", "") or "").lower()
    data = file_storage.read()
    if name.endswith(".csv"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        text = data.decode("utf-8-sig", errors="replace")
        for r in csv.reader(io.StringIO(text)):
            ws.append(r)
        return wb
    return load_workbook(io.BytesIO(data))


def _sheet_matches(rule_sheet, sheet_name):
    return rule_sheet in ("*", None) or rule_sheet == sheet_name


# ----- main entry ----------------------------------------------------------
def validate_workbook(file_storage, rules=None):
    """Validate a workbook. Returns (xlsx_bytes, summary_dict).

    summary_dict = {total_rows, checked_cells, fail, warn, failures:[...]}
    """
    rules = (rules or DEFAULT_RULES).get("rules", [])
    wb = _load_workbook(file_storage)

    failures = []
    total_rows = 0

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = _s(ws.cell(1, c).value)
            if h:
                headers[h.lower()] = c
        for r in range(2, ws.max_row + 1):
            total_rows += 1
            row = {}
            for hlow, c in headers.items():
                row[hlow] = ws.cell(r, c).value
            for rule in rules:
                if not _sheet_matches(rule.get("sheet"), ws.title):
                    continue
                col = (rule.get("column") or "").lower()
                fn = CHECKS.get(rule.get("check"))
                if not fn or col not in headers:
                    continue
                cell_val = ws.cell(r, headers[col]).value
                sev, msg = fn(cell_val, row, rule)
                if sev:
                    cell = ws.cell(r, headers[col])
                    cell.fill = FILL_FAIL if sev == SEV_FAIL else FILL_WARN
                    failures.append({
                        "sheet": ws.title, "row": r, "column": rule.get("column"),
                        "value": _s(cell_val), "severity": sev,
                        "rule": rule.get("check"), "message": msg,
                    })

    _append_summary(wb, failures, total_rows)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    n_fail = sum(1 for f in failures if f["severity"] == SEV_FAIL)
    n_warn = sum(1 for f in failures if f["severity"] == SEV_WARN)
    summary = {
        "total_rows": total_rows,
        "issues": len(failures),
        "fail": n_fail,
        "warn": n_warn,
        "failures": failures[:500],
    }
    return out.getvalue(), summary


def _append_summary(wb, failures, total_rows):
    if "Validation Summary" in wb.sheetnames:
        del wb["Validation Summary"]
    ws = wb.create_sheet("Validation Summary", 0)
    n_fail = sum(1 for f in failures if f["severity"] == SEV_FAIL)
    n_warn = sum(1 for f in failures if f["severity"] == SEV_WARN)
    ws.append(["Validation Summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Rows checked", total_rows])
    ws.append(["Failures (red)", n_fail])
    ws.append(["Warnings (amber)", n_warn])
    ws.append([])
    head = ["Sheet", "Row", "Column", "Value", "Severity", "Rule", "Message"]
    ws.append(head)
    hr = ws.max_row
    for c in range(1, len(head) + 1):
        cell = ws.cell(hr, c)
        cell.fill = FILL_HEAD
        cell.font = Font(bold=True, color="FFFFFFFF")
    for f in failures:
        ws.append([f["sheet"], f["row"], f["column"], f["value"][:200],
                   f["severity"], f["rule"], f["message"]])
        ws.cell(ws.max_row, 5).fill = FILL_FAIL if f["severity"] == SEV_FAIL else FILL_WARN
    widths = [16, 6, 22, 40, 10, 30, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for c in range(1, len(head) + 1):
        ws.cell(hr, c).alignment = Alignment(vertical="center")
    return ws
